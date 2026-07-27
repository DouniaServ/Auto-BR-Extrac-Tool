from __future__ import annotations

import os
import base64
import io
import json
import time
import uuid
import hashlib
import traceback
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict, Any

import pandas as pd
import streamlit as st
import openpyxl  # needed to preserve VBA from template.xlsm
from openpyxl.styles import Alignment, Font, PatternFill

# Optional AG Grid
HAS_AGGRID = False
# AG Grid disabled in this stable version to avoid Streamlit component errors.
# The review table uses Streamlit's native st.data_editor with larger height options.
HAS_AGGRID = False
try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode, JsCode  # optional, not used here
except Exception:
    AgGrid = GridOptionsBuilder = GridUpdateMode = DataReturnMode = JsCode = None


# =============================================================================
# Tool version (V1.0)
# =============================================================================
TOOL_VERSION = "1.8.2"

# =============================================================================
# Extractor loader (keeps each site's logic untouched)
# =============================================================================
def get_extractor(site: str):
    if site == "Arklow":
        from sites import arklow
        return arklow.extract
    if site == "Anpharm":
        from sites import anpharm
        return anpharm.extract
    if site == "Gidy":
        from sites import gidy
        return gidy.extract
    if site == "Toledo":
        from sites import toledo
        return toledo.extract
    if site == "Bolbec":
        from sites import bolbec
        return bolbec.extract
    raise ValueError(f"No extractor implemented for site: {site}")


# =============================================================================
# Per-site schema
# =============================================================================
SITES = ["Arklow", "Anpharm", "Gidy", "Toledo", "Bolbec"]

REVIEW_OK_COL = "Reviewed & OK"
REVIEWER_COL = "Reviewer"
REVIEWED_AT_COL = "Reviewed at"
REVIEWER_NOTE_COL = "Reviewer note"

# Internal columns used only inside Streamlit review workspace.
# They are removed automatically from final Excel exports.
INTERNAL_ROW_ID_COL = "__review_row_id__"

# Row action used inside the Streamlit review grid.
# This replaces the old Insert above / Insert below checkbox columns with one cleaner Excel-like action column.
ROW_ACTION_COL = "__row_action__"
ROW_ACTION_NONE = ""
ROW_ACTION_INSERT_ABOVE = "Insert row above"
ROW_ACTION_INSERT_BELOW = "Insert row below"
ROW_ACTION_DELETE = "Delete row"
ROW_ACTION_CHOICES = [
    ROW_ACTION_NONE,
    ROW_ACTION_INSERT_ABOVE,
    ROW_ACTION_INSERT_BELOW,
    ROW_ACTION_DELETE,
]

# Legacy temporary columns from previous iterations. Keep them here only so old saved sessions/export clean-up stays safe.
INSERT_ABOVE_COL = "__insert_above__"
INSERT_BELOW_COL = "__insert_below__"
REMOVE_ROW_COL = "__remove_row__"

# Temporary UI/action columns used only by the Streamlit review grid.
# They are never exported to Excel and are reset after use.
TEMP_UI_COLS = {ROW_ACTION_COL, INSERT_ABOVE_COL, INSERT_BELOW_COL, REMOVE_ROW_COL}
INTERNAL_COLS = {INTERNAL_ROW_ID_COL, *TEMP_UI_COLS}

# Server-side autosave folder. Review data is persisted outside Streamlit session_state
# so a browser refresh / laptop lock / websocket reset does not lose reviewer progress.
REVIEW_STORAGE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "review_sessions")


def ensure_internal_row_ids(df: pd.DataFrame) -> pd.DataFrame:
    """Add a stable internal row ID so audit trail remains clear after filtering/sorting."""
    if not isinstance(df, pd.DataFrame):
        return df

    out = df.copy()
    if INTERNAL_ROW_ID_COL not in out.columns:
        out.insert(0, INTERNAL_ROW_ID_COL, [str(uuid.uuid4())[:8] for _ in range(len(out))])
    else:
        mask = out[INTERNAL_ROW_ID_COL].isna() | (out[INTERNAL_ROW_ID_COL].astype(str).str.strip() == "")
        if mask.any():
            out.loc[mask, INTERNAL_ROW_ID_COL] = [str(uuid.uuid4())[:8] for _ in range(int(mask.sum()))]
    return out


def drop_internal_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Remove Streamlit-only control columns before validation/export/display where needed."""
    if not isinstance(df, pd.DataFrame):
        return df
    return df.drop(columns=[c for c in INTERNAL_COLS if c in df.columns], errors="ignore")


def get_schema_columns_for_export(df: pd.DataFrame, schema: Dict[str, Any]) -> List[str]:
    """Return final business columns only; internal review columns are never exported."""
    cols = derive_columns_from_schema_or_df(drop_internal_cols(df), schema)
    return [c for c in cols if c not in INTERNAL_COLS]

SITE_SCHEMA: Dict[str, Dict[str, Any]] = {
    "Arklow": {
        "columns": [
            "PDF name",
            "Data number",
            "Page number",
            "Paragraph number",
            "Process step",
            "Sub process step",
            "Data title",
            "Tag",
            REVIEW_OK_COL,
            REVIEWER_COL,
            REVIEWED_AT_COL,
            REVIEWER_NOTE_COL,
        ],
        "required_nonempty": ["Data number", "Page number", "Process step", "Data title"],
        "page_col": "Page number",
        "page_min": 1,
        "dup_key": ["Page number", "Paragraph number", "Data title", "Tag"],
        "search_cols": ["Process step", "Sub process step", "Data title", "Tag"],
        "review_ok_col": REVIEW_OK_COL,
        "reviewer_col": REVIEWER_COL,
        "reviewed_at_col": REVIEWED_AT_COL,
        "review_note_col": REVIEWER_NOTE_COL,
    },
    "Anpharm": {
        "columns": [
            "PDF name",
            "Data number",
            "Page number",
            "Paragraph number",
            "Process step",
            "Sub process step",
            "Data Title",
            "Data tag",
            REVIEW_OK_COL,
            REVIEWER_COL,
            REVIEWED_AT_COL,
            REVIEWER_NOTE_COL,
        ],
        "required_nonempty": ["Data number", "Page number", "Process step", "Data Title"],
        "page_col": "Page number",
        "page_min": 1,
        "dup_key": ["Page number", "Paragraph number", "Data Title", "Value", "Data tag"],
        "search_cols": ["Process step", "Sub process step", "Data Title", "Value", "Data tag"],
        "review_ok_col": REVIEW_OK_COL,
        "reviewer_col": REVIEWER_COL,
        "reviewed_at_col": REVIEWED_AT_COL,
        "review_note_col": REVIEWER_NOTE_COL,
    },
    "Gidy": {
        "columns": [
            "PDF name",
            "Data number",
            "Page number",
            "Paragraph number",
            "Process step",
            "Sub process step",
            "Data Title",
            "Data tag",
            REVIEW_OK_COL,
            REVIEWER_COL,
            REVIEWED_AT_COL,
            REVIEWER_NOTE_COL,
        ],
        "required_nonempty": ["Data number", "Page number", "Process step", "Data Title"],
        "page_col": "Page number",
        "page_min": 1,
        "dup_key": ["Page number", "Paragraph number", "Data Title", "Data tag"],
        "search_cols": ["Process step", "Sub process step", "Data Title", "Data tag", "Théorique", "Réel", "Visa"],
        "review_ok_col": REVIEW_OK_COL,
        "reviewer_col": REVIEWER_COL,
        "reviewed_at_col": REVIEWED_AT_COL,
        "review_note_col": REVIEWER_NOTE_COL,
    },
    "Toledo": {
        "columns": [],
        "required_nonempty": [],
        "page_col": None,
        "page_min": 1,
        "dup_key": [],
        "search_cols": [],
        "review_ok_col": REVIEW_OK_COL,
        "reviewer_col": REVIEWER_COL,
        "reviewed_at_col": REVIEWED_AT_COL,
        "review_note_col": REVIEWER_NOTE_COL,
    },
    "Bolbec": {
        "columns": [],
        "required_nonempty": [],
        "page_col": None,
        "page_min": 1,
        "dup_key": [],
        "search_cols": [],
        "review_ok_col": REVIEW_OK_COL,
        "reviewer_col": REVIEWER_COL,
        "reviewed_at_col": REVIEWED_AT_COL,
        "review_note_col": REVIEWER_NOTE_COL,
    },
}

REVIEW_STATUS_CHOICES = ["Draft", "In review", "Approved", "Rejected"]


# =============================================================================
# UI Config + CSS
# =============================================================================
st.set_page_config(page_title="Auto BR Extractor", page_icon="📄", layout="wide")


def inject_background_css(image_name="background image 2.jpg"):
    paths = [
        image_name,
        os.path.join(os.getcwd(), image_name),
        os.path.join(os.path.dirname(__file__), image_name),
    ]
    real_path = next((p for p in paths if os.path.exists(p)), None)

    if real_path:
        with open(real_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        ext = os.path.splitext(real_path)[1].lower()
        mime = "png" if ext == ".png" else "jpeg"
        bg = f'url("data:image/{mime};base64,{b64}")'
    else:
        bg = "linear-gradient(135deg, #020b1f, #08162b)"

    st.markdown(
        f"""
        <style>
        html, body, .stApp,
        [data-testid="stAppViewContainer"],
        [data-testid="stAppViewContainer"] > .main {{
            background: {bg} !important;
            background-size: cover !important;
            background-position: center center !important;
            background-repeat: no-repeat !important;
            background-attachment: fixed !important;
        }}
        section.main {{ background: transparent !important; }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def inject_ui_css():
    st.markdown(
        """
        <style>
        html, body, [class*="css"], .stMarkdown, .stText, .stMarkdown p {
          color: #ffffff !important;
          font-family: "Inter", Arial, sans-serif !important;
        }
        .stTabs [data-baseweb="tab"] {
          font-family: Arial, sans-serif !important;
          font-weight: 900 !important;
          color: #ffffff !important;
          font-size: 18px !important;
        }
        .stTabs [data-baseweb="tab"][aria-selected="true"] { color: #ffffff !important; }

        div[data-testid="stWidgetLabel"],
        div[data-testid="stWidgetLabel"] *,
        label,
        .stTextInput label,
        .stTextArea label,
        .stSelectbox label,
        .stNumberInput label,
        .stFileUploader label,
        .stMultiSelect label,
        .stRadio label,
        .stCheckbox label,
        [data-baseweb="form-control-label"],
        [data-baseweb="form-control-label"] * {
          color: #ffffff !important;
          opacity: 1 !important;
          font-weight: 900 !important;
        }

        h3,
        div[data-testid="stSubheader"],
        div[data-testid="stSubheader"] *,
        h2 {
          color: #ffffff !important;
          font-weight: 1000 !important;
          opacity: 1 !important;
        }

        input, textarea { color: #08162b !important; font-weight: 900 !important; }

        .stCaption, div[data-testid="stCaptionContainer"] * {
          color: #ffffff !important;
          opacity: 0.95 !important;
          font-weight: 800 !important;
        }

        .kebab-btn button{
          border-radius: 12px !important;
          padding: 6px 12px !important;
          font-size: 22px !important;
          font-weight: 900 !important;
          background: rgba(255,255,255,0.14) !important;
          border: 1px solid rgba(255,255,255,0.35) !important;
          color: #ffffff !important;
          box-shadow: 0 10px 24px rgba(0,0,0,0.35) !important;
        }
        .kebab-btn button:hover{ background: rgba(255,255,255,0.22) !important; }

        div[data-testid="stExpander"]{
          background: rgba(10, 18, 35, 0.62) !important;
          border: 1px solid rgba(255,255,255,0.18) !important;
          border-radius: 16px !important;
          box-shadow: 0 10px 28px rgba(0,0,0,0.35) !important;
          backdrop-filter: blur(10px) !important;
        }
        div[data-testid="stExpander"] *{ color: #ffffff !important; font-weight: 800 !important; }

        .metric-card{
          background: rgba(10, 18, 35, 0.62);
          border-radius: 16px;
          padding: 18px;
          border: 1px solid rgba(255,255,255,0.18);
          box-shadow: 0 10px 28px rgba(0,0,0,0.35);
          backdrop-filter: blur(10px);
        }
        .metric-label{
          font-size: 13px;
          font-weight: 900;
          letter-spacing: 0.3px;
          color: #ffffff !important;
          margin-bottom: 6px;
          text-transform: uppercase;
        }
        .metric-value{
          font-size: 22px;
          font-weight: 900;
          color: #ffffff !important;
          word-break: break-word;
        }
        .hr{ height: 1px; background: rgba(255,255,255,0.22); margin: 12px 0 14px; }

        div[data-baseweb="select"] > div{
          background: #ffffff !important;
          border: 2px solid rgba(255,255,255,0.92) !important;
          border-radius: 14px !important;
          min-height: 54px !important;
          box-shadow: 0 10px 22px rgba(0,0,0,0.25) !important;
        }
        div[data-baseweb="select"] span{
          color: #08162b !important;
          font-size: 18px !important;
          font-weight: 900 !important;
        }
        div[data-baseweb="select"] svg{ color: #08162b !important; }
        div[data-baseweb="menu"]{ background: #ffffff !important; border-radius: 14px !important; }

        .panel{
          background: rgba(10, 18, 35, 0.62);
          border: 1px solid rgba(255,255,255,0.18);
          border-radius: 16px;
          padding: 16px;
          box-shadow: 0 10px 28px rgba(0,0,0,0.35);
          backdrop-filter: blur(10px);
        }

        div[data-testid="stDataFrame"],
        div[data-testid="stDataFrame"] * { color: #08162b !important; }

        /* Native Streamlit table: dark border frame.
           Goal: avoid the flat all-white-paper effect without adding colored
           row markers. Streamlit's native st.data_editor is canvas-based, so
           editable cell background colors cannot be fully controlled. This
           keeps the editor stable and adds a dark, visible table frame. */
        div[data-testid="stDataEditor"] {
          position: relative !important;
          background: #111827 !important;              /* dark table frame */
          border: 4px solid #0f172a !important;        /* dark outer border */
          border-radius: 14px !important;
          padding: 8px !important;
          box-shadow: 0 16px 36px rgba(0,0,0,0.42) !important;
          overflow: hidden !important;
        }

        /* Inner dark separator around the editable grid */
        div[data-testid="stDataEditor"] > div {
          position: relative !important;
          border: 2px solid #334155 !important;
          border-radius: 10px !important;
          overflow: hidden !important;
          background: #1f2937 !important;
        }

        /* Neutral canvas frame. The cells may stay white in Streamlit native editor,
           but the table is now clearly framed and segmented by a dark border. */
        div[data-testid="stDataEditor"] canvas {
          border: 1px solid #475569 !important;
          border-radius: 8px !important;
          background-color: #f8fafc !important;
        }

        /* Header/cell enhancement when Streamlit exposes DOM grid roles. */
        div[data-testid="stDataEditor"] [role="grid"],
        div[data-testid="stDataEditor"] [role="table"] {
          border: 2px solid #0f172a !important;
          border-collapse: collapse !important;
          background: #f8fafc !important;
        }
        div[data-testid="stDataEditor"] [role="columnheader"] {
          background: #cbd5e1 !important;
          border-right: 1px solid #334155 !important;
          border-bottom: 2px solid #0f172a !important;
          color: #08162b !important;
          font-weight: 900 !important;
        }
        div[data-testid="stDataEditor"] [role="gridcell"] {
          border-right: 1px solid #64748b !important;
          border-bottom: 1px solid #64748b !important;
          background: #f8fafc !important;
          color: #08162b !important;
          font-weight: 650 !important;
        }
        div[data-testid="stDataEditor"] * {
          color: #08162b !important;
          font-weight: 650 !important;
        }

        .footer-version{
          position: fixed;
          bottom: 14px;
          left: 0;
          width: 100%;
          text-align: center;
          opacity: 0.95;
          font-size: 20px;
          font-weight: 1000;
          letter-spacing: 0.5px;
          color: #ffffff;
          z-index: 9999;
          pointer-events: none;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def step_title(step_key: str, text: str):
    active = st.session_state.view_step == step_key
    color = "#ffffff" if active else "rgba(255,255,255,0.75)"
    st.markdown(
        f"<h2 style='color:{color}; font-weight:1000; margin-bottom:16px;'>{text}</h2>",
        unsafe_allow_html=True,
    )


inject_background_css("background image 2.jpg")
inject_ui_css()


def inject_review_fullscreen_css():
    """Make the review workspace occupy the full Streamlit viewport.

    This removes Streamlit chrome/padding and uses a true 100vw layout.
    A separate browser Full Screen button is also rendered so reviewers can
    switch the browser itself to fullscreen when supported.
    """
    st.markdown(
        """
        <style>
        [data-testid="stHeader"],
        [data-testid="stToolbar"],
        [data-testid="stDecoration"],
        header,
        footer {
            display: none !important;
            visibility: hidden !important;
            height: 0px !important;
            min-height: 0px !important;
        }

        html, body, .stApp,
        [data-testid="stAppViewContainer"],
        [data-testid="stAppViewContainer"] > .main {
            width: 100vw !important;
            min-width: 100vw !important;
            max-width: 100vw !important;
            height: 100vh !important;
            min-height: 100vh !important;
            overflow: hidden !important;
        }

        .block-container {
            max-width: 100vw !important;
            width: 100vw !important;
            padding-top: 0rem !important;
            padding-bottom: 0rem !important;
            padding-left: 0.25rem !important;
            padding-right: 0.25rem !important;
        }

        section.main > div {
            padding-top: 0rem !important;
            padding-bottom: 0rem !important;
        }

        /* In focus mode, the review page should look like a dedicated app screen. */
        .panel {
            width: calc(100vw - 12px) !important;
            min-height: calc(100vh - 8px) !important;
            margin: 0 !important;
            border-radius: 0 !important;
            padding: 8px !important;
            background: rgba(6, 12, 28, 0.96) !important;
            border: 0 !important;
            box-shadow: none !important;
        }

        .fullscreen-review-banner {
            position: sticky;
            top: 0;
            z-index: 999999;
            background: rgba(10, 18, 35, 0.96);
            border: 1px solid rgba(255,255,255,0.20);
            border-radius: 0;
            padding: 8px 12px;
            margin: 0 0 6px 0;
            box-shadow: 0 8px 22px rgba(0,0,0,0.35);
        }

        .fullscreen-toolbar-note {
            background: rgba(255,255,255,0.08);
            border: 1px solid rgba(255,255,255,0.18);
            border-radius: 10px;
            padding: 6px 10px;
            margin: 4px 0 8px 0;
            color: #ffffff !important;
            font-weight: 800;
        }

        /* Make AG Grid consume space better in focus mode. */
        .review-grid-shell {
            border-radius: 6px !important;
            padding: 4px !important;
            margin-bottom: 0 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_fullscreen_instruction():
    """Safe full-screen guidance without using Streamlit custom HTML components.

    The previous version used `components.html()` to trigger browser fullscreen.
    Some Streamlit environments block that iframe/JS component and show a component error.
    This version keeps full-screen as an app-level focus mode and asks the reviewer
    to use the browser/OS shortcut for real full-screen.
    """
    st.info(
        "Full-screen mode is now component-free. For real browser full-screen, press F11 "
        "or use your browser full-screen command. The app still hides the main header/navigation "
        "and gives a larger review table."
    )


def inject_review_grid_readability_css():
    """Small wrapper styling around the review grid.

    The real row segmentation is handled by AG Grid custom_css when available.
    This wrapper makes the review workspace look like a controlled spreadsheet
    instead of a flat white block.
    """
    st.markdown(
        """
        <style>
        .review-grid-shell {
            background: #111827;
            border: 4px solid #0f172a;
            border-radius: 14px;
            padding: 8px;
            box-shadow: 0 16px 36px rgba(0,0,0,0.42);
        }
        .review-grid-caption {
            color: #ffffff !important;
            font-weight: 900;
            margin-bottom: 8px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def get_aggrid_custom_css() -> Dict[str, Dict[str, str]]:
    """Clear spreadsheet-like segmentation for AG Grid rows/cells."""
    return {
        ".ag-root-wrapper": {
            "border": "2px solid #94a3b8 !important",
            "border-radius": "12px !important",
            "overflow": "hidden !important",
            "background-color": "#ffffff !important",
        },
        ".ag-header": {
            "background-color": "#e2e8f0 !important",
            "border-bottom": "2px solid #94a3b8 !important",
        },
        ".ag-header-cell": {
            "background-color": "#e2e8f0 !important",
            "color": "#0f172a !important",
            "font-weight": "900 !important",
            "font-size": "14px !important",
            "border-right": "1px solid #94a3b8 !important",
        },
        ".ag-row": {
            "border-bottom": "1px solid #cbd5e1 !important",
        },
        ".ag-row-even": {
            "background-color": "#ffffff !important",
        },
        ".ag-row-odd": {
            "background-color": "#f1f5f9 !important",
        },
        ".ag-row-hover": {
            "background-color": "#dbeafe !important",
        },
        ".ag-cell": {
            "color": "#0f172a !important",
            "font-size": "14px !important",
            "font-weight": "650 !important",
            "border-right": "1px solid #e2e8f0 !important",
            "padding-left": "10px !important",
            "padding-right": "10px !important",
            "display": "flex !important",
            "align-items": "center !important",
            "line-height": "1.25 !important",
        },
        ".ag-cell-focus": {
            "border": "2px solid #2563eb !important",
            "background-color": "#eff6ff !important",
        },
        ".ag-pinned-left-cols-container .ag-cell": {
            "background-color": "#eef2ff !important",
            "font-weight": "900 !important",
        },
        ".ag-checkbox-input-wrapper": {
            "transform": "scale(1.15)",
        },
        ".ag-cell input[type='checkbox']": {
            "accent-color": "#2563eb !important",
        },
    }


def get_aggrid_review_ok_checkbox_renderer():
    """
    Real checkbox renderer for AG Grid.

    Some streamlit-aggrid versions do not render agCheckboxCellRenderer as a visible checkbox.
    This custom renderer creates a native checkbox and writes the boolean value back to the grid data.
    """
    return JsCode(
        """
        function(params) {
            const checkbox = document.createElement('input');
            checkbox.type = 'checkbox';
            checkbox.checked = (params.value === true || params.value === 'true' || params.value === 1 || params.value === '1');
            checkbox.style.width = '18px';
            checkbox.style.height = '18px';
            checkbox.style.cursor = 'pointer';
            checkbox.style.margin = 'auto';
            checkbox.disabled = params.colDef.editable === false;

            checkbox.addEventListener('click', function(event) {
                event.stopPropagation();
            });

            checkbox.addEventListener('change', function(event) {
                params.node.setDataValue(params.colDef.field, event.target.checked);
            });

            return checkbox;
        }
        """
    )


# =============================================================================
# NEW: Arklow-only metadata fill (Product name -> Process step)
# =============================================================================
def apply_ui_metadata_to_df(site: str, df: pd.DataFrame, file_name: str) -> pd.DataFrame:
    """
    Sites using UI Product name for Process step:
      - Arklow
      - Bolbec
      - Toledo

    Also sets file name in first row if column exists.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df

    # Only apply to selected sites
    if site not in {"Arklow", "Bolbec", "Toledo"}:
        return df

    product = (st.session_state.product_name or "").strip() or "--"

    # Process step override
    if "Process step" in df.columns:
        df["Process step"] = product

    # File name in first row
    if "Unnamed: 0" in df.columns:
        df["Unnamed: 0"] = None
        df.loc[df.index[0], "Unnamed: 0"] = file_name

    if "PDF name" in df.columns:
        df["PDF name"] = None
        df.loc[df.index[0], "PDF name"] = file_name

    return df


# =============================================================================
# State helpers
# =============================================================================
def get_active_schema(site: str) -> Dict[str, Any]:
    return SITE_SCHEMA.get(site, {}) or {}


def derive_columns_from_schema_or_df(df: Optional[pd.DataFrame], schema: Dict[str, Any]) -> List[str]:
    cols = schema.get("columns") or []
    if cols:
        return cols
    if isinstance(df, pd.DataFrame):
        return list(df.columns)
    return []


def derive_search_cols(schema: Dict[str, Any], cols: List[str]) -> List[str]:
    sc = schema.get("search_cols") or []
    if sc:
        return [c for c in sc if c in cols]
    return cols


def ss_init():
    ss = st.session_state
    ss.setdefault("job_id", str(uuid.uuid4())[:8])
    ss.setdefault("site", SITES[0])
    ss.setdefault("product_name", "")

    ss.setdefault("batch_record_bytes", None)
    ss.setdefault("batch_record_name", None)
    ss.setdefault("batch_record_fp", None)

    ss.setdefault("df", None)

    # Timing / KPI
    ss.setdefault("extraction_started_at", "")
    ss.setdefault("extraction_ended_at", "")
    ss.setdefault("extraction_seconds", None)      # float seconds
    ss.setdefault("extraction_rows", 0)           # int
    ss.setdefault("extraction_source", "")        # "pdf_word" | "review_pack"

    # Validation (user-action only)
    ss.setdefault("validation", None)
    ss.setdefault("user_validated", False)
    ss.setdefault("validation_status", "Not validated yet")

    # keep only search + view_mode
    ss.setdefault("search", "")
    ss.setdefault("view_mode", "Not reviewed only")

    ss.setdefault("audit_log", [])
    ss.setdefault("show_menu", False)

    ss.setdefault("reviewer_name", os.getenv("REVIEWER_DEFAULT", ""))
    ss.setdefault("review_status", "Draft")
    ss.setdefault("review_comment", "")
    ss.setdefault("reviewed_at", "")
    ss.setdefault("baseline_hash", "")

    ss.setdefault("skipped_pages", [])
    ss.setdefault("view_step", "Setup")

    ss.setdefault("blocking_msgs", [])

    ss.setdefault("active_schema", get_active_schema(ss.site))
    ss.setdefault("active_columns", derive_columns_from_schema_or_df(ss.df, ss.active_schema))
    ss.setdefault("extraction_control_df", pd.DataFrame())
    ss.setdefault("extraction_control_pdf_bytes", None)
    ss.setdefault("extraction_control_pdf_marker", "")
    ss.setdefault("last_saved_at", "")
    ss.setdefault("full_screen_review", False)
    ss.setdefault("review_table_size_mode", "Large")
    ss.setdefault("review_table_height_normal", 850)
    ss.setdefault("review_table_height_fullscreen", 1050)


def audit(event: str, details: str = ""):
    st.session_state.audit_log.append(
        {
            "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
            "job_id": st.session_state.job_id,
            "event": event,
            "details": details,
        }
    )


def reset_all():
    st.session_state.clear()
    ss_init()


ss_init()


# =============================================================================
# Guards (steps always enabled; clicking warns)
# =============================================================================
def needs_product() -> bool:
    return not bool((st.session_state.product_name or "").strip())


def needs_data() -> bool:
    return not isinstance(st.session_state.df, pd.DataFrame)


def set_blocking_messages(msgs: List[str]):
    st.session_state.blocking_msgs = msgs


def go_or_warn(target_step: str):
    msgs = []
    if needs_product():
        msgs.append("• Please enter **Product name** first (Step 1).")
    if needs_data():
        msgs.append("• Please **upload Batch Record (PDF/Word)** OR **upload Review Pack (Excel)** first (Step 1).")

    if msgs:
        set_blocking_messages(msgs)
        st.session_state.view_step = "Setup"
        return

    set_blocking_messages([])
    st.session_state.view_step = target_step


def enforce_step_guard_soft():
    """
    Safety net if someone ends up in Review/Export without prerequisites.
    """
    if st.session_state.view_step in ("Review", "Export Excel"):
        msgs = []
        if needs_product():
            msgs.append("• Please enter **Product name** first (Step 1).")
        if needs_data():
            msgs.append("• Please **upload Batch Record (PDF/Word)** OR **upload Review Pack (Excel)** first (Step 1).")

        if msgs:
            set_blocking_messages(msgs)
            st.session_state.view_step = "Setup"
            st.stop()

def style_checklist_column(ws, header_name=REVIEW_OK_COL):
    checklist_col = None

    for cell in ws[1]:
        if str(cell.value).strip() == header_name:
            checklist_col = cell.column
            break

    if checklist_col is None:
        return

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=checklist_col)

        value = str(cell.value).strip().lower()

        if value in {"true", "1", "yes", "y", "checked", "☑", "☒", "✓", "✔", "x", "þ"}:
            cell.value = "þ"   # checked checkbox in Wingdings
        else:
            cell.value = "¨"   # unchecked checkbox in Wingdings

        cell.font = Font(name="Wingdings", size=16, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
# =============================================================================
# Hash + dirty
# =============================================================================
def df_content_hash(df: pd.DataFrame) -> str:
    # Internal row IDs should not change the business hash.
    clean_df = drop_internal_cols(df)
    b = clean_df.to_csv(index=False).encode("utf-8")
    return hashlib.sha256(b).hexdigest()[:16]


def mark_dirty(reason: str):
    st.session_state.user_validated = False
    st.session_state.validation_status = "Not validated yet"

    if st.session_state.review_status == "Approved":
        st.session_state.review_status = "Draft"
        st.session_state.reviewed_at = ""
        audit("approval_revoked_on_edit", reason)


# =============================================================================
# Validation
# =============================================================================
@dataclass
class ValidationResult:
    ok: bool
    errors: pd.DataFrame
    warnings: pd.DataFrame
    flagged_row_ids: List[int]


def validate_df(df: pd.DataFrame, schema: Dict[str, Any]) -> ValidationResult:
    df = drop_internal_cols(df)
    cols = derive_columns_from_schema_or_df(df, schema)
    required = schema.get("required_nonempty") or []
    page_col = schema.get("page_col")
    page_min = int(schema.get("page_min", 1))
    dup_key = schema.get("dup_key") or []

    errors: List[dict] = []
    warnings: List[dict] = []
    flagged = set()

    # schema cols missing
    missing_cols = [c for c in cols if c not in df.columns]
    if missing_cols:
        for c in missing_cols:
            errors.append({"type": "schema", "row": None, "column": c, "message": f"Missing column: {c}"})
        return ValidationResult(False, pd.DataFrame(errors), pd.DataFrame(warnings), [])

    # required non-empty
    for col in required:
        if col not in df.columns:
            errors.append({"type": "schema", "row": None, "column": col, "message": f"Missing required column: {col}"})
            continue
        empty = df[col].isna() | (df[col].astype(str).str.strip() == "")
        for i in df.index[empty].tolist()[:5000]:
            errors.append({"type": "required", "row": int(i), "column": col, "message": "Value required"})
            flagged.add(int(i))

    # page checks
    if page_col and page_col in df.columns:
        p = pd.to_numeric(df[page_col], errors="coerce")
        bad_page = p.isna() | (p < page_min) | (p % 1 != 0)
        for i in df.index[bad_page].tolist()[:5000]:
            warnings.append(
                {"type": "page", "row": int(i), "column": page_col, "message": f"Page must be integer >= {page_min}"}
            )
            flagged.add(int(i))

    # duplicates (warning) ✅ V1.0 implemented
    if dup_key:
        usable = [c for c in dup_key if c in df.columns]
        if usable:
            key_df = df[usable].copy()

            nonempty_mask = True
            for c in usable:
                nonempty_mask = nonempty_mask & (~key_df[c].isna()) & (key_df[c].astype(str).str.strip() != "")
            if isinstance(nonempty_mask, bool):
                nonempty_mask = pd.Series([True] * len(df), index=df.index)

            dups = df.loc[nonempty_mask].duplicated(subset=usable, keep=False)
            dup_idx = df.loc[nonempty_mask].index[dups].tolist()[:5000]
            for i in dup_idx:
                warnings.append(
                    {
                        "type": "duplicate",
                        "row": int(i),
                        "column": ",".join(usable),
                        "message": f"Possible duplicate row based on key: {usable}",
                    }
                )
                flagged.add(int(i))

    err_df = pd.DataFrame(errors)
    warn_df = pd.DataFrame(warnings)
    ok = len(err_df) == 0
    return ValidationResult(ok, err_df, warn_df, sorted(flagged))

# =============================================================================
# Extraction control helpers
# =============================================================================
def _first_existing_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None

def build_extraction_control_df(ui_df: pd.DataFrame) -> pd.DataFrame:
    """
    Fallback only. Real Full Source Traceability must come from extraction_control.py
    because it needs the original PDF/raw text.
    """
    return pd.DataFrame(columns=[
        "Page",
        "Source ID",
        "Raw PDF text",
        "Source group",
        "Reconciliation status",
        "Matched Extracted reference",
        "Reviewer note",
    ])


def build_extraction_control_summary_df(control_df: pd.DataFrame) -> pd.DataFrame:
    """
    Simple Quality summary for the Full Source Traceability sheet.
    The sheet is traceability support only; the Extracted sheet remains the main review output.
    """
    if control_df is None or control_df.empty:
        return pd.DataFrame(columns=["Metric", "Value"])

    status_col = "Reconciliation status" if "Reconciliation status" in control_df.columns else "Match status"
    status = control_df[status_col].fillna("").astype(str).str.strip()

    source_group = (
        control_df["Source group"].fillna("").astype(str).str.strip()
        if "Source group" in control_df.columns
        else pd.Series([""] * len(control_df))
    )

    covered = (
        status.eq("Reconciled with structured pre-draft")
        | status.eq("Reconciled with structured pre-draft")
        | status.eq("Covered")
    )
    metadata = status.eq("Document metadata / context") | source_group.eq("Document metadata / context")
    visible_not_reconciled = (
        status.eq("Visible in source - not reconciled")
        | status.eq("Not found in Extracted")
        | status.eq("Not covered")
    ) & source_group.eq("Source data")

    return pd.DataFrame([
        {"Metric": "Total readable source items visible", "Value": int(len(control_df))},
        {"Metric": "Reconciled with structured pre-draft", "Value": int(covered.sum())},
        {"Metric": "Document metadata / context", "Value": int(metadata.sum())},
        {"Metric": "Visible source data not reconciled", "Value": int(visible_not_reconciled.sum())},
        {"Metric": "Quality position", "Value": "Pre-draft support: main review remains Extracted sheet against the PDF."},
    ])


def build_potential_missing_data_df(control_df: pd.DataFrame) -> pd.DataFrame:
    """
    Optional support sheet: only source-data rows not reconciled with the Extracted sheet.
    This avoids asking reviewers to review the full traceability matrix line by line.
    """
    cols = [
        "Page",
        "Source ID",
        "Raw PDF text",
        "Source group",
        "Reconciliation status",
        "Matched Extracted reference",
        "Reviewer note",
    ]

    if not isinstance(control_df, pd.DataFrame) or control_df.empty:
        return pd.DataFrame(columns=cols)

    for col in cols:
        if col not in control_df.columns:
            control_df[col] = ""

    mask = (
        control_df["Source group"].fillna("").astype(str).str.strip().eq("Source data")
        & control_df["Reconciliation status"].fillna("").astype(str).str.strip().isin([
            "Visible in source - not reconciled",
            "Not found in Extracted",
            "Not covered",
        ])
    )

    return control_df.loc[mask, cols].copy()


def build_extracted_row_control_df(extracted_df: pd.DataFrame, control_df: pd.DataFrame) -> pd.DataFrame:
    """
    Reverse quality control:
    verifies that each extracted row has at least one supporting raw PDF item.
    """
    if not isinstance(extracted_df, pd.DataFrame) or extracted_df.empty:
        return pd.DataFrame(columns=[
            "Extracted row",
            "Page number",
            "Data title",
            "Tag",
            "Support status",
            "Supporting raw text count",
            "Reviewer action",
        ])

    if not isinstance(control_df, pd.DataFrame) or control_df.empty:
        return pd.DataFrame(columns=[
            "Extracted row",
            "Page number",
            "Data title",
            "Tag",
            "Support status",
            "Supporting raw text count",
            "Reviewer action",
        ])

    matched_col = "Matched extracted row"
    if matched_col not in control_df.columns:
        return pd.DataFrame()

    supported_rows = (
        pd.to_numeric(control_df[matched_col], errors="coerce")
        .dropna()
        .astype(int)
        .value_counts()
        .to_dict()
    )

    rows = []

    for idx, row in extracted_df.iterrows():
        support_count = int(supported_rows.get(int(idx), 0))

        if support_count > 0:
            status = "Supported by raw PDF"
            action = "OK"
        else:
            status = "No raw PDF support found"
            action = "Review extracted row"

        rows.append({
            "Extracted row": int(idx),
            "Page number": row.get("Page number", ""),
            "Data title": row.get("Data title", row.get("Data Title", "")),
            "Tag": row.get("Tag", row.get("Data tag", "")),
            "Support status": status,
            "Supporting raw text count": support_count,
            "Reviewer action": action,
        })

    return pd.DataFrame(rows)


# =============================================================================
# Extraction Control PDF report
# =============================================================================
def _pdf_safe_text(value: Any, max_chars: int = 600) -> str:
    """Convert any cell value into PDF-safe, readable text."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    text = str(value).replace("\r", " ").replace("\n", " ").replace("\t", " ").strip()
    text = " ".join(text.split())
    if len(text) > max_chars:
        return text[: max_chars - 3] + "..."
    return text


def build_extraction_control_pdf_report(
    control_df: pd.DataFrame,
    extracted_df: Optional[pd.DataFrame] = None,
    review_df: Optional[pd.DataFrame] = None,
    audit_df: Optional[pd.DataFrame] = None,
) -> bytes:
    """
    Build a Quality-friendly PDF report from the Full Source Traceability / Extraction Control data.

    The Excel export remains the detailed controlled output. This PDF is a readable report with:
    - review metadata
    - extraction control summary
    - rows requiring reviewer attention
    - full traceability appendix with raw text shortened for PDF readability
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception as e:
        raise ImportError(
            "PDF report generation requires the 'reportlab' package. "
            "Please add reportlab to requirements.txt."
        ) from e

    if not isinstance(control_df, pd.DataFrame) or control_df.empty:
        control_df = build_extraction_control_df(extracted_df if isinstance(extracted_df, pd.DataFrame) else pd.DataFrame())

    summary_df = build_extraction_control_summary_df(control_df)
    potential_missing_df = build_potential_missing_data_df(control_df)

    buffer = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=0.7 * cm,
        leftMargin=0.7 * cm,
        topMargin=0.7 * cm,
        bottomMargin=0.9 * cm,
        title="Extraction Control Report",
        author=st.session_state.get("reviewer_name", ""),
    )

    base_styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=base_styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0f172a"),
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=base_styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1f2937"),
        spaceBefore=8,
        spaceAfter=6,
    )
    normal_style = ParagraphStyle(
        "ReportNormal",
        parent=base_styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#111827"),
        spaceAfter=4,
    )
    note_style = ParagraphStyle(
        "ReportNote",
        parent=normal_style,
        fontName="Helvetica-Oblique",
        textColor=colors.HexColor("#334155"),
    )
    cell_style = ParagraphStyle(
        "TableCell",
        parent=base_styles["Normal"],
        fontName="Helvetica",
        fontSize=6.4,
        leading=7.4,
        textColor=colors.HexColor("#111827"),
    )
    header_cell_style = ParagraphStyle(
        "TableHeaderCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
        fontSize=6.8,
        leading=7.8,
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    def _table(data: List[List[Any]], col_widths: Optional[List[float]] = None, repeat_rows: int = 1) -> Table:
        t = Table(data, colWidths=col_widths, repeatRows=repeat_rows, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#94a3b8")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        return t

    story: List[Any] = []
    story.append(Paragraph("Extraction Control Report", title_style))
    story.append(Paragraph("Pre-draft support report - human review remains mandatory before use.", normal_style))

    generated_at = time.strftime("%Y-%m-%d %H:%M:%S")
    metadata_rows = [
        [Paragraph("Field", header_cell_style), Paragraph("Value", header_cell_style)],
        [Paragraph("Site", cell_style), Paragraph(_pdf_safe_text(st.session_state.get("site", ""), 120), cell_style)],
        [Paragraph("Product", cell_style), Paragraph(_pdf_safe_text(st.session_state.get("product_name", ""), 120), cell_style)],
        [Paragraph("Batch Record file", cell_style), Paragraph(_pdf_safe_text(st.session_state.get("batch_record_name", ""), 200), cell_style)],
        [Paragraph("Job ID", cell_style), Paragraph(_pdf_safe_text(st.session_state.get("job_id", ""), 80), cell_style)],
        [Paragraph("Reviewer", cell_style), Paragraph(_pdf_safe_text(st.session_state.get("reviewer_name", ""), 120), cell_style)],
        [Paragraph("Review status", cell_style), Paragraph(_pdf_safe_text(st.session_state.get("review_status", ""), 120), cell_style)],
        [Paragraph("Reviewed at", cell_style), Paragraph(_pdf_safe_text(st.session_state.get("reviewed_at", ""), 120), cell_style)],
        [Paragraph("Generated at", cell_style), Paragraph(generated_at, cell_style)],
        [Paragraph("Tool version", cell_style), Paragraph(_pdf_safe_text(TOOL_VERSION, 50), cell_style)],
        [Paragraph("Extracted rows", cell_style), Paragraph(str(len(extracted_df)) if isinstance(extracted_df, pd.DataFrame) else "-", cell_style)],
        [Paragraph("Control rows", cell_style), Paragraph(str(len(control_df)), cell_style)],
    ]
    story.append(Spacer(1, 0.15 * cm))
    story.append(_table(metadata_rows, col_widths=[4.0 * cm, 22.5 * cm], repeat_rows=1))

    story.append(Paragraph("Quality position", subtitle_style))
    story.append(Paragraph(
        "This report supports source traceability for the AI pre-draft extraction. It does not replace the mandatory business review against the Batch Record template.",
        note_style,
    ))

    if isinstance(summary_df, pd.DataFrame) and not summary_df.empty:
        story.append(Paragraph("Extraction control summary", subtitle_style))
        summary_rows = [[Paragraph("Metric", header_cell_style), Paragraph("Value", header_cell_style)]]
        for _, row in summary_df.iterrows():
            summary_rows.append([
                Paragraph(_pdf_safe_text(row.get("Metric", ""), 220), cell_style),
                Paragraph(_pdf_safe_text(row.get("Value", ""), 300), cell_style),
            ])
        story.append(_table(summary_rows, col_widths=[9.0 * cm, 17.5 * cm], repeat_rows=1))

    if isinstance(potential_missing_df, pd.DataFrame) and not potential_missing_df.empty:
        story.append(Paragraph("Rows requiring reviewer attention", subtitle_style))
        story.append(Paragraph(
            "These rows are visible source-data items that were not reconciled with the structured pre-draft.",
            note_style,
        ))
        attention_cols = [
            "Page",
            "Source ID",
            "Reconciliation status",
            "Matched Extracted reference",
            "Raw PDF text",
            "Reviewer note",
        ]
        for col in attention_cols:
            if col not in potential_missing_df.columns:
                potential_missing_df[col] = ""
        attention_rows = [[Paragraph(c, header_cell_style) for c in attention_cols]]
        for _, row in potential_missing_df[attention_cols].iterrows():
            attention_rows.append([
                Paragraph(_pdf_safe_text(row.get("Page", ""), 40), cell_style),
                Paragraph(_pdf_safe_text(row.get("Source ID", ""), 80), cell_style),
                Paragraph(_pdf_safe_text(row.get("Reconciliation status", ""), 130), cell_style),
                Paragraph(_pdf_safe_text(row.get("Matched Extracted reference", ""), 150), cell_style),
                Paragraph(_pdf_safe_text(row.get("Raw PDF text", ""), 500), cell_style),
                Paragraph(_pdf_safe_text(row.get("Reviewer note", ""), 200), cell_style),
            ])
        story.append(_table(attention_rows, col_widths=[1.4 * cm, 2.2 * cm, 4.2 * cm, 4.2 * cm, 11.0 * cm, 3.5 * cm], repeat_rows=1))
    else:
        story.append(Paragraph("Rows requiring reviewer attention", subtitle_style))
        story.append(Paragraph("No unreconciled source-data rows identified in the extraction control data.", normal_style))

    story.append(PageBreak())
    story.append(Paragraph("Full Source Traceability appendix", subtitle_style))
    story.append(Paragraph(
        "For PDF readability, long raw text cells are shortened in this report. The final Excel export remains the detailed controlled output.",
        note_style,
    ))

    appendix_cols = [
        "Page",
        "Source ID",
        "Source group",
        "Reconciliation status",
        "Matched Extracted reference",
        "Reviewer note",
        "Raw PDF text",
    ]
    for col in appendix_cols:
        if col not in control_df.columns:
            control_df[col] = ""

    col_widths = [1.2 * cm, 1.8 * cm, 3.0 * cm, 4.0 * cm, 4.0 * cm, 3.2 * cm, 9.0 * cm]
    chunk_size = 80
    for chunk_start in range(0, len(control_df), chunk_size):
        chunk = control_df.iloc[chunk_start: chunk_start + chunk_size]
        appendix_rows = [[Paragraph(c, header_cell_style) for c in appendix_cols]]
        for _, row in chunk[appendix_cols].iterrows():
            appendix_rows.append([
                Paragraph(_pdf_safe_text(row.get("Page", ""), 35), cell_style),
                Paragraph(_pdf_safe_text(row.get("Source ID", ""), 60), cell_style),
                Paragraph(_pdf_safe_text(row.get("Source group", ""), 90), cell_style),
                Paragraph(_pdf_safe_text(row.get("Reconciliation status", ""), 130), cell_style),
                Paragraph(_pdf_safe_text(row.get("Matched Extracted reference", ""), 130), cell_style),
                Paragraph(_pdf_safe_text(row.get("Reviewer note", ""), 150), cell_style),
                Paragraph(_pdf_safe_text(row.get("Raw PDF text", ""), 500), cell_style),
            ])
        story.append(_table(appendix_rows, col_widths=col_widths, repeat_rows=1))
        if chunk_start + chunk_size < len(control_df):
            story.append(PageBreak())

    def _footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#475569"))
        footer_text = f"Extraction Control Report - Job ID: {st.session_state.get('job_id', '')} - Page {doc_obj.page}"
        canvas.drawString(0.7 * cm, 0.35 * cm, footer_text)
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _dataframe_pdf_fingerprint(df: Optional[pd.DataFrame]) -> str:
    """Stable fingerprint used to refresh the cached PDF only when its content changes."""
    if not isinstance(df, pd.DataFrame):
        return "no-dataframe"
    try:
        payload = df.to_csv(index=False).encode("utf-8")
    except Exception:
        payload = repr(df).encode("utf-8", errors="replace")
    return hashlib.sha256(payload).hexdigest()[:16]


def render_extraction_control_pdf_download(location_key: str = "step2") -> None:
    """Render the Extraction Control PDF download inside the controlled review step.

    The report is available while the review is Draft, In review or Approved. It reflects
    the last data applied/saved in the controlled table.
    """
    extracted_df = drop_internal_cols(st.session_state.df)
    control_df = (
        st.session_state.extraction_control_df.copy()
        if isinstance(st.session_state.extraction_control_df, pd.DataFrame)
        and not st.session_state.extraction_control_df.empty
        else build_extraction_control_df(extracted_df)
    )
    review_df = make_review_sheet_df()

    metadata_marker = json.dumps(
        {
            "site": st.session_state.get("site", ""),
            "product_name": st.session_state.get("product_name", ""),
            "batch_record_name": st.session_state.get("batch_record_name", ""),
            "job_id": st.session_state.get("job_id", ""),
            "reviewer_name": st.session_state.get("reviewer_name", ""),
            "review_status": st.session_state.get("review_status", ""),
            "reviewed_at": st.session_state.get("reviewed_at", ""),
            "review_comment": st.session_state.get("review_comment", ""),
            "tool_version": TOOL_VERSION,
        },
        sort_keys=True,
        ensure_ascii=True,
    )
    marker = hashlib.sha256(
        (
            _dataframe_pdf_fingerprint(extracted_df)
            + _dataframe_pdf_fingerprint(control_df)
            + _dataframe_pdf_fingerprint(review_df)
            + metadata_marker
        ).encode("utf-8")
    ).hexdigest()[:20]

    st.subheader("Extraction Control PDF")
    st.caption(
        "Download the current Extraction Control / Full Source Traceability report during Step 2. "
        "The PDF reflects the last changes applied with Apply / Save. Approval is not required."
    )

    try:
        if (
            st.session_state.get("extraction_control_pdf_marker") != marker
            or not isinstance(st.session_state.get("extraction_control_pdf_bytes"), (bytes, bytearray))
        ):
            with st.spinner("Preparing Extraction Control PDF report…"):
                st.session_state.extraction_control_pdf_bytes = build_extraction_control_pdf_report(
                    control_df=control_df,
                    extracted_df=extracted_df,
                    review_df=review_df,
                    audit_df=None,
                )
                st.session_state.extraction_control_pdf_marker = marker

        control_pdf_filename = (
            f"{st.session_state.site}_Extraction_Control_Report_"
            f"{time.strftime('%Y%m%d_%H%M')}.pdf"
        )
        downloaded = st.download_button(
            label="📄 Download Extraction Control PDF",
            data=st.session_state.extraction_control_pdf_bytes,
            file_name=control_pdf_filename,
            mime="application/pdf",
            use_container_width=True,
            key=f"download_extraction_control_pdf_{location_key}",
        )
        if downloaded:
            audit("extraction_control_pdf_downloaded", control_pdf_filename)
            autosave_review_session("extraction_control_pdf_downloaded")

    except ImportError:
        st.error(
            "PDF report generation requires the reportlab package. "
            "Add reportlab to requirements.txt, then restart the app."
        )
        st.code("reportlab==4.2.2")
    except Exception as e:
        st.error(f"Could not prepare the Extraction Control PDF report: {type(e).__name__}: {e}")


# =============================================================================
# Excel helpers (review sheet, parse, diff)
# =============================================================================
def make_review_sheet_df() -> pd.DataFrame:
    skipped = st.session_state.skipped_pages or []
    skipped_txt = ",".join(map(str, skipped)) if skipped else ""

    sec = st.session_state.extraction_seconds
    sec_txt = "" if sec is None else f"{sec:.2f}"
    rows = int(st.session_state.extraction_rows or 0)
    rate_txt = ""
    if sec is not None and sec > 0 and rows > 0:
        rate_txt = f"{(rows / sec):.2f} rows/s"

    return pd.DataFrame(
        [
            {"Field": "Site", "Value": st.session_state.site},
            {"Field": "Product", "Value": st.session_state.product_name},
            {"Field": "File name", "Value": st.session_state.batch_record_name or ""},
            {"Field": "Job ID", "Value": st.session_state.job_id},
            {"Field": "Reviewer", "Value": st.session_state.reviewer_name},
            {"Field": "Status", "Value": st.session_state.review_status},
            {"Field": "Reviewed at", "Value": st.session_state.reviewed_at},
            {"Field": "Total rows extracted", "Value": len(st.session_state.df) if st.session_state.df is not None else 0},
            {"Field": "Skipped pages", "Value": skipped_txt},
            {"Field": "Validation status", "Value": st.session_state.validation_status},
            {"Field": "Baseline hash", "Value": st.session_state.baseline_hash},
            {"Field": "Reviewer comment", "Value": st.session_state.review_comment},
            {"Field": "Extraction source", "Value": st.session_state.extraction_source},
            {"Field": "Extraction started at", "Value": st.session_state.extraction_started_at},
            {"Field": "Extraction ended at", "Value": st.session_state.extraction_ended_at},
            {"Field": "Extraction duration (s)", "Value": sec_txt},
            {"Field": "Extraction throughput", "Value": rate_txt},
        ]
    )


def parse_review_sheet(review_df: Optional[pd.DataFrame]) -> Dict[str, str]:
    if review_df is None or len(review_df) == 0:
        return {}
    cols = [c.lower().strip() for c in review_df.columns]
    if "field" not in cols or "value" not in cols:
        return {}
    field_col = review_df.columns[cols.index("field")]
    value_col = review_df.columns[cols.index("value")]
    out = {}
    for _, r in review_df.iterrows():
        k = str(r[field_col]).strip()
        v = "" if pd.isna(r[value_col]) else str(r[value_col])
        out[k] = v
    return out


def diff_summary(old: pd.DataFrame, new: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    out = []
    n = min(len(old), len(new))
    old2 = old.iloc[:n].reset_index(drop=True)
    new2 = new.iloc[:n].reset_index(drop=True)
    for col in cols:
        o = old2[col].astype(str) if col in old2.columns else pd.Series([""] * n)
        nn = new2[col].astype(str) if col in new2.columns else pd.Series([""] * n)
        out.append({"column": col, "changed_cells": int((o != nn).sum())})
    out.append({"column": "__row_count__", "changed_cells": int(abs(len(old) - len(new)))})
    return pd.DataFrame(out)


# =============================================================================
# Robust Excel sheet name handling
# =============================================================================
def _norm_sheet_name(s: str) -> str:
    return str(s).strip().lower()


def find_sheet_name(xl: pd.ExcelFile, desired: str) -> Optional[str]:
    """Exact match ignoring case and surrounding spaces."""
    desired_norm = _norm_sheet_name(desired)
    for s in xl.sheet_names:
        if _norm_sheet_name(s) == desired_norm:
            return s
    return None


def find_sheet_by_contains(xl: pd.ExcelFile, token: str) -> Optional[str]:
    """Fallback: first sheet name that contains token (case-insensitive)."""
    t = token.strip().lower()
    for s in xl.sheet_names:
        if t in _norm_sheet_name(s):
            return s
    return None


# =============================================================================
# Workflow columns normalization
# =============================================================================
REVIEW_OK_DEFAULTS = {
    REVIEW_OK_COL: False,
    REVIEWER_COL: "",
    REVIEWED_AT_COL: "",
    REVIEWER_NOTE_COL: "",
}

ALIASES_TO_CANONICAL = {
    "Needs review": "__OLD_PENDING__",
    "Pending review": "__OLD_PENDING__",
    "Pending Review": "__OLD_PENDING__",
    "Needs Review": "__OLD_PENDING__",
    "Reviewed and OK": REVIEW_OK_COL,
    "Reviewed OK": REVIEW_OK_COL,
    "Reviewed At": REVIEWED_AT_COL,
    "Review note": REVIEWER_NOTE_COL,
    "Reviewer Note": REVIEWER_NOTE_COL,
}


def coerce_review_ok_series(s: pd.Series) -> pd.Series:
    """
    Normalize Review OK values for the app checkbox.
    Important: bool("☐") is True in Python, so never use astype(bool)
    directly on this column.
    """
    def one(v):
        if pd.isna(v):
            return False
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)):
            return bool(v)
        t = str(v).strip().lower()
        if t in {"", "false", "0", "no", "n", "unchecked", "☐", "□", "none", "nan"}:
            return False
        if t in {"true", "1", "yes", "y", "checked", "☑", "☒", "✓", "✔", "x"}:
            return True
        # Unknown text should default to False to avoid marking rows reviewed by mistake.
        return False

    return s.map(one).astype(bool)



def review_ok_to_excel_symbols(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    if REVIEW_OK_COL in out.columns:
        ok = coerce_review_ok_series(out[REVIEW_OK_COL])
        out[REVIEW_OK_COL] = ok.map(lambda x: "þ" if x else "¨")

    return out

def ensure_workflow_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns={c: str(c).strip() for c in df.columns})
    rename_map = {c: ALIASES_TO_CANONICAL[c] for c in df.columns if c in ALIASES_TO_CANONICAL}
    if rename_map:
        df = df.rename(columns=rename_map)

    if "__OLD_PENDING__" in df.columns and REVIEW_OK_COL not in df.columns:
        old_pending = coerce_review_ok_series(df["__OLD_PENDING__"])
        df[REVIEW_OK_COL] = (~old_pending)
        df = df.drop(columns=["__OLD_PENDING__"])

    for col, default in REVIEW_OK_DEFAULTS.items():
        if col not in df.columns:
            df[col] = default

    df[REVIEW_OK_COL] = coerce_review_ok_series(df[REVIEW_OK_COL])
    return df


# =============================================================================
# Macro-enabled Excel export
# =============================================================================
# Path to the macro-enabled template (.xlsm). Created once, kept next to app.py.
# The template must contain three empty sheets ("Extracted", "Review", "Audit log")
# and the VBA Worksheet_BeforeDoubleClick on the "Extracted" sheet that
# toggles ☑/☐ and stamps the date in "Reviewed at".
MACRO_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "template.xlsm",
)

EXT_XLSM = "xlsm"
EXT_XLSX = "xlsx"
MIME_XLSM = "application/vnd.ms-excel.sheet.macroEnabled.12"
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _macro_template_available() -> bool:
    return os.path.exists(MACRO_TEMPLATE_PATH)


def _clear_worksheet(ws) -> None:
    """Remove all rows but keep the sheet — preserves its VBA code module binding."""
    if ws.max_row and ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)


def _write_df_to_worksheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to an openpyxl worksheet starting at A1."""
    if df is None:
        return

    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))

    if df.empty:
        return

    for i, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for j, val in enumerate(row, start=1):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                ws.cell(row=i, column=j, value=None)
            else:
                ws.cell(row=i, column=j, value=val)

    style_checklist_column(ws, REVIEW_OK_COL)

def build_excel_bytes_macro(
    extracted_df: pd.DataFrame,
    review_df: pd.DataFrame,
    audit_df: Optional[pd.DataFrame],
    include_audit: bool = True,
    control_df_override: Optional[pd.DataFrame] = None,
) -> Tuple[bytes, str, str]:
    """
    Build the export workbook.

    Quality-facing design:
    - Extracted = main pre-draft review sheet
    - Review = reviewer/status metadata
    - Full Source Traceability = simple raw-source traceability support
    - Audit log = user/action traceability
    """
    export_df = review_ok_to_excel_symbols(drop_internal_cols(extracted_df))

    if isinstance(control_df_override, pd.DataFrame):
        control_df = control_df_override.copy()
    else:
        control_df = (
            st.session_state.extraction_control_df.copy()
            if hasattr(st.session_state, "extraction_control_df")
            and isinstance(st.session_state.extraction_control_df, pd.DataFrame)
            else pd.DataFrame()
        )

    required_control_cols = {
        "Page",
        "Source ID",
        "Raw PDF text",
        "Source group",
        "Reconciliation status",
        "Matched Extracted reference",
        "Reviewer note",
    }
    ordered_control_cols = [
        "Page",
        "Source ID",
        "Raw PDF text",
        "Source group",
        "Reconciliation status",
        "Matched Extracted reference",
        "Reviewer note",
    ]

    if control_df.empty or not required_control_cols.issubset(set(control_df.columns)):
        audit(
            "invalid_full_source_traceability_export",
            f"Full Source Traceability is empty or invalid format. Columns found: {list(control_df.columns)}"
        )
        control_df = build_extraction_control_df(extracted_df)

    for col in ordered_control_cols:
        if col not in control_df.columns:
            control_df[col] = ""

    control_df = control_df[ordered_control_cols]
    if _macro_template_available():
        try:
            wb = openpyxl.load_workbook(MACRO_TEMPLATE_PATH, keep_vba=True)

            # Normalize old control sheet names from previous versions.
            target_sheet = "Full Source Traceability"
            legacy_candidates = [
                "Full Source Traceability",
                "Full PDF Raw Extraction",
                "Extraction control",
                "Extraction Control",
            ]

            if target_sheet not in wb.sheetnames:
                renamed = False
                for legacy_sheet in legacy_candidates:
                    if legacy_sheet in wb.sheetnames:
                        wb[legacy_sheet].title = target_sheet
                        renamed = True
                        break
                if not renamed:
                    wb.create_sheet(target_sheet)

            for legacy_sheet in legacy_candidates:
                if legacy_sheet in wb.sheetnames and legacy_sheet != target_sheet:
                    wb.remove(wb[legacy_sheet])

            # Remove the optional Potential Missing Data sheet from official review pack.
            # The pre-draft review remains focused on Extracted + PDF;
            # Full Source Traceability stays as support evidence only.
            if "Potential Missing Data" in wb.sheetnames:
                wb.remove(wb["Potential Missing Data"])

            for sheet_name, df_to_write in [
                ("Extracted", export_df),
                ("Review", review_df),
                (target_sheet, control_df),
            ]:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)

                ws = wb[sheet_name]
                _clear_worksheet(ws)
                _write_df_to_worksheet(ws, df_to_write)

            if include_audit:
                if "Audit log" not in wb.sheetnames:
                    wb.create_sheet("Audit log")

                ws = wb["Audit log"]
                _clear_worksheet(ws)
                _write_df_to_worksheet(
                    ws,
                    audit_df if audit_df is not None else pd.DataFrame()
                )

            out = io.BytesIO()
            wb.save(out)
            return out.getvalue(), MIME_XLSM, EXT_XLSM

        except Exception as e:
            audit("macro_template_failed_fallback_xlsx", f"{type(e).__name__}: {e}")

    out = io.BytesIO()

    with pd.ExcelWriter(out, engine="openpyxl") as w:
        export_df.to_excel(w, index=False, sheet_name="Extracted")
        review_df.to_excel(w, index=False, sheet_name="Review")
        control_df.to_excel(w, index=False, sheet_name="Full Source Traceability")

        if include_audit:
            (audit_df if audit_df is not None else pd.DataFrame()).to_excel(
                w,
                index=False,
                sheet_name="Audit log"
            )

    return out.getvalue(), MIME_XLSX, EXT_XLSX

def df_to_excel_bytes(df: pd.DataFrame, include_audit: bool = True) -> Tuple[bytes, str, str]:
    """
    Thin wrapper used by the review-pack download in Step 2.
    Returns (file_bytes, mime_type, extension).
    """
    propagate_reviewer_to_df(fill_all_rows=False)
    review_df = make_review_sheet_df()
    audit_df = pd.DataFrame(st.session_state.audit_log) if include_audit else None

    return build_excel_bytes_macro(
        extracted_df=df,
        review_df=review_df,
        audit_df=audit_df,
        include_audit=include_audit,
    )

# =============================================================================
# Extraction plumbing
# =============================================================================
def get_extractor_control_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Get Full Source Traceability control from df.attrs.
    """
    if not isinstance(df, pd.DataFrame):
        return pd.DataFrame()

    required_cols = {
        "Page",
        "Source ID",
        "Raw PDF text",
        "Source group",
        "Reconciliation status",
    }

    for key in ("extraction_control", "Extraction control", "Extraction Control"):
        control = getattr(df, "attrs", {}).get(key)

        if isinstance(control, pd.DataFrame) and not control.empty:
            if required_cols.issubset(set(control.columns)):
                return control.copy()

    return build_extraction_control_df(df)

def fp(b: bytes) -> str:
    return f"{len(b)}:{hash(b[:2048])}"


def extract_by_site(site: str, file_bytes: bytes, file_name: str) -> pd.DataFrame:
    extractor = get_extractor(site)
    df = extractor(file_bytes, file_name)

    # Add missing review/workflow columns
    df = ensure_workflow_columns(df)

    # Fill UI metadata for Arklow / Bolbec / Toledo
    df = apply_ui_metadata_to_df(site, df, file_name)

    # =========================================================
    # ALWAYS rebuild Full Source Traceability from the PDF.
    # This keeps all readable raw PDF text visible.
    # No skipped-by-design / not expected logic.
    # =========================================================
    full_raw_df = pd.DataFrame()

    try:
        import tempfile

        tmp_pdf_path = None

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(file_bytes)
                tmp_pdf_path = tmp.name

            from extraction_control import build_strong_raw_order_control

            full_raw_df = build_strong_raw_order_control(
                tmp_pdf_path,
                df,
                threshold_extracted=0.85,
                threshold_partial=0.40,
                include_summary=False,
            )

        finally:
            if tmp_pdf_path and os.path.exists(tmp_pdf_path):
                try:
                    os.remove(tmp_pdf_path)
                except Exception:
                    pass

    except Exception as e:
        audit("full_raw_extraction_failed", f"{site}: {type(e).__name__}: {e}")
        full_raw_df = build_extraction_control_df(df)

    # =========================================================
    # Apply site schema columns to Extracted sheet only
    # =========================================================
    schema = get_active_schema(site)
    cols = schema.get("columns") or list(df.columns)

    if schema.get("columns"):
        missing = [c for c in cols if c not in df.columns]
        if missing:
            raise ValueError(f"Extractor for {site} returned missing columns: {missing}")

        df = df[cols].copy()

    # Reattach Full Source Traceability AFTER df copy/schema trim
    df.attrs["extraction_control"] = full_raw_df

    return df


def maybe_auto_extract():
    ss = st.session_state
    if ss.batch_record_bytes is None:
        return

    new_fp = fp(ss.batch_record_bytes)

    if ss.batch_record_fp == new_fp and isinstance(ss.df, pd.DataFrame):
        ss.df = apply_ui_metadata_to_df(ss.site, ss.df, ss.batch_record_name or "record")
        return

    ss.extraction_source = "pdf_word"
    ss.extraction_started_at = time.strftime("%Y-%m-%d %H:%M:%S")
    t0 = time.perf_counter()

    schema = get_active_schema(ss.site)
    df = extract_by_site(ss.site, ss.batch_record_bytes, ss.batch_record_name or "record")

    t1 = time.perf_counter()
    ss.extraction_ended_at = time.strftime("%Y-%m-%d %H:%M:%S")
    ss.extraction_seconds = float(t1 - t0)
    ss.extraction_rows = int(len(df))

    ss.df = ensure_internal_row_ids(df)
    ss.extraction_control_df = get_extractor_control_from_df(df)
    ss.active_schema = schema
    ss.active_columns = derive_columns_from_schema_or_df(ss.df, ss.active_schema)

    ss.validation = validate_df(ss.df, ss.active_schema)
    ss.user_validated = False
    ss.validation_status = "Not validated yet"

    ss.batch_record_fp = new_fp

    ss.baseline_hash = df_content_hash(ss.df)
    ss.review_status = "Draft"
    ss.reviewed_at = ""
    ss.review_comment = ""

    audit(
        "auto_extracted",
        f"site={ss.site}; rows={len(ss.df)}; file={ss.batch_record_name}; seconds={ss.extraction_seconds:.2f}",
    )


def load_review_pack_excel(file_bytes: bytes, expected_cols: List[str]) -> Tuple[pd.DataFrame, Optional[pd.DataFrame]]:
    bio = io.BytesIO(file_bytes)
    xl = pd.ExcelFile(bio, engine="openpyxl")

    extracted_sheet = find_sheet_name(xl, "Extracted") or find_sheet_by_contains(xl, "extract")
    if not extracted_sheet:
        raise ValueError(
            "Missing sheet 'Extracted' in uploaded Excel. "
            f"Found sheets: {xl.sheet_names}. "
            "Please upload the review pack exported from this tool."
        )

    review_sheet = find_sheet_name(xl, "Review") or find_sheet_by_contains(xl, "review")

    extracted = pd.read_excel(xl, extracted_sheet, engine="openpyxl")
    review = pd.read_excel(xl, review_sheet, engine="openpyxl") if review_sheet else None

    extracted = ensure_workflow_columns(extracted)

    missing = [c for c in expected_cols if c not in extracted.columns]
    if missing:
        raise ValueError(f"Uploaded Excel missing required columns: {missing}")

    return extracted[expected_cols].copy(), review


def resume_from_review_pack(file_bytes: bytes, file_name: str):
    bio = io.BytesIO(file_bytes)
    xl = pd.ExcelFile(bio, engine="openpyxl")

    extracted_sheet = find_sheet_name(xl, "Extracted") or find_sheet_by_contains(xl, "extract")
    if not extracted_sheet:
        raise ValueError(
            "Missing sheet 'Extracted' in uploaded Excel. "
            f"Found sheets: {xl.sheet_names}. "
            "Please upload the review pack exported from this tool."
        )

    review_sheet = find_sheet_name(xl, "Review") or find_sheet_by_contains(xl, "review")
    control_sheet = (
        find_sheet_name(xl, "Full Source Traceability")
        or find_sheet_name(xl, "Full Source Coverage Matrix")
        or find_sheet_name(xl, "Full PDF Raw Extraction")
        or find_sheet_name(xl, "Extraction control")
        or find_sheet_name(xl, "Extraction Control")
        or find_sheet_by_contains(xl, "raw")
        or find_sheet_by_contains(xl, "control")
    )

    extracted = pd.read_excel(xl, extracted_sheet, engine="openpyxl")
    review_df = pd.read_excel(xl, review_sheet, engine="openpyxl") if review_sheet else None
    control_df = pd.read_excel(xl, control_sheet, engine="openpyxl") if control_sheet else pd.DataFrame()
    meta = parse_review_sheet(review_df)

    site = meta.get("Site", "") or st.session_state.site
    if site not in SITES:
        site = st.session_state.site

    st.session_state.site = site
    st.session_state.product_name = meta.get("Product", "") or st.session_state.product_name
    st.session_state.batch_record_name = meta.get("File name", "") or file_name

    st.session_state.active_schema = get_active_schema(site)
    expected_cols = derive_columns_from_schema_or_df(None, st.session_state.active_schema)

    extracted = ensure_workflow_columns(extracted)

    extracted = apply_ui_metadata_to_df(site, extracted, st.session_state.batch_record_name or file_name)

    missing = [c for c in expected_cols if c not in extracted.columns]
    if missing:
        raise ValueError(f"Uploaded Excel missing required columns for site '{site}': {missing}")

    st.session_state.df = ensure_internal_row_ids(extracted[expected_cols].copy())
    st.session_state.extraction_control_df = (
        control_df if isinstance(control_df, pd.DataFrame) and not control_df.empty
        else build_extraction_control_df(st.session_state.df)
    )
    st.session_state.active_columns = expected_cols

    st.session_state.validation = validate_df(st.session_state.df, st.session_state.active_schema)
    st.session_state.user_validated = False
    st.session_state.validation_status = "Not validated yet"

    st.session_state.baseline_hash = df_content_hash(st.session_state.df)

    st.session_state.review_status = "In review"
    st.session_state.reviewed_at = meta.get("Reviewed at", "") or ""
    st.session_state.review_comment = meta.get("Reviewer comment", "") or ""
    st.session_state.reviewer_name = meta.get("Reviewer", "") or st.session_state.reviewer_name

    st.session_state.extraction_source = meta.get("Extraction source", "") or "review_pack"
    st.session_state.extraction_started_at = meta.get("Extraction started at", "") or ""
    st.session_state.extraction_ended_at = meta.get("Extraction ended at", "") or ""
    try:
        sec = meta.get("Extraction duration (s)", "")
        st.session_state.extraction_seconds = float(sec) if str(sec).strip() else None
    except Exception:
        st.session_state.extraction_seconds = None

    st.session_state.extraction_rows = int(len(st.session_state.df))

    if needs_product():
        set_blocking_messages(["• Please enter **Product name** first (Step 1)."])
        st.session_state.view_step = "Setup"
        audit("resumed_from_review_pack_missing_product", f"file={file_name}; site={site}; rows={len(st.session_state.df)}")
    else:
        set_blocking_messages([])
        st.session_state.view_step = "Review"
        audit("resumed_from_review_pack", f"file={file_name}; site={site}; rows={len(st.session_state.df)}")

    autosave_review_session("legacy_excel_migrated")


# =============================================================================
# Server-side autosave / resume
# =============================================================================
def _safe_job_id(job_id: str) -> str:
    """Keep folder names safe and deterministic."""
    safe = "".join(ch for ch in str(job_id) if ch.isalnum() or ch in {"-", "_"}).strip()
    return safe or str(uuid.uuid4())[:8]


def get_review_session_folder(job_id: str) -> str:
    folder = os.path.join(REVIEW_STORAGE_DIR, _safe_job_id(job_id))
    os.makedirs(folder, exist_ok=True)
    return folder


def _write_json(path: str, payload: Any) -> None:
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False, default=str)
    os.replace(tmp_path, path)


def _read_json(path: str, default: Any) -> Any:
    try:
        if not os.path.exists(path):
            return default
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


def _df_to_records(df: Optional[pd.DataFrame]) -> List[Dict[str, Any]]:
    if not isinstance(df, pd.DataFrame):
        return []

    safe_df = df.copy()
    # JSON cannot represent NaN cleanly for later review; convert to empty string.
    safe_df = safe_df.where(pd.notna(safe_df), "")
    return safe_df.to_dict(orient="records")


def _records_to_df(records: Any) -> pd.DataFrame:
    if not isinstance(records, list):
        return pd.DataFrame()
    return pd.DataFrame(records)


def build_review_metadata() -> Dict[str, Any]:
    df = st.session_state.get("df")
    ok_col = REVIEW_OK_COL
    reviewed_count = 0
    total_rows = int(len(df)) if isinstance(df, pd.DataFrame) else 0
    if isinstance(df, pd.DataFrame) and ok_col in df.columns:
        reviewed_count = int(coerce_review_ok_series(df[ok_col]).sum())

    return {
        "job_id": st.session_state.get("job_id", ""),
        "site": st.session_state.get("site", ""),
        "product_name": st.session_state.get("product_name", ""),
        "batch_record_name": st.session_state.get("batch_record_name", ""),
        "batch_record_fp": st.session_state.get("batch_record_fp", ""),
        "reviewer_name": st.session_state.get("reviewer_name", ""),
        "review_status": st.session_state.get("review_status", "Draft"),
        "reviewed_at": st.session_state.get("reviewed_at", ""),
        "review_comment": st.session_state.get("review_comment", ""),
        "validation_status": st.session_state.get("validation_status", "Not validated yet"),
        "baseline_hash": st.session_state.get("baseline_hash", ""),
        "extraction_source": st.session_state.get("extraction_source", ""),
        "extraction_started_at": st.session_state.get("extraction_started_at", ""),
        "extraction_ended_at": st.session_state.get("extraction_ended_at", ""),
        "extraction_seconds": st.session_state.get("extraction_seconds", None),
        "extraction_rows": st.session_state.get("extraction_rows", total_rows),
        "total_rows": total_rows,
        "reviewed_rows": reviewed_count,
        "not_reviewed_rows": max(total_rows - reviewed_count, 0),
        "last_saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "tool_version": TOOL_VERSION,
    }


def autosave_review_session(reason: str = "autosave") -> bool:
    """Persist current review state outside Streamlit session_state.

    This is the control that protects reviewers from losing progress after:
    - browser refresh
    - laptop lock / sleep
    - network disconnect
    - Streamlit websocket reset
    - app rerun
    """
    df = st.session_state.get("df")
    if not isinstance(df, pd.DataFrame):
        return False

    job_id = st.session_state.get("job_id") or str(uuid.uuid4())[:8]
    st.session_state.job_id = _safe_job_id(job_id)
    folder = get_review_session_folder(st.session_state.job_id)

    try:
        metadata = build_review_metadata()
        metadata["autosave_reason"] = reason

        _write_json(os.path.join(folder, "metadata.json"), metadata)
        _write_json(os.path.join(folder, "current_review.json"), _df_to_records(ensure_internal_row_ids(df)))
        _write_json(os.path.join(folder, "audit_log.json"), st.session_state.get("audit_log", []))

        control_df = st.session_state.get("extraction_control_df")
        _write_json(os.path.join(folder, "extraction_control.json"), _df_to_records(control_df))

        st.session_state.last_saved_at = metadata["last_saved_at"]
        return True
    except Exception as e:
        st.warning(f"Autosave failed: {type(e).__name__}: {e}")
        return False


def list_saved_review_sessions() -> List[Dict[str, Any]]:
    if not os.path.exists(REVIEW_STORAGE_DIR):
        return []

    sessions: List[Dict[str, Any]] = []
    for job_id in os.listdir(REVIEW_STORAGE_DIR):
        folder = os.path.join(REVIEW_STORAGE_DIR, job_id)
        if not os.path.isdir(folder):
            continue
        meta = _read_json(os.path.join(folder, "metadata.json"), {})
        if isinstance(meta, dict) and meta.get("job_id"):
            sessions.append(meta)

    sessions.sort(key=lambda x: str(x.get("last_saved_at", "")), reverse=True)
    return sessions


def load_saved_review_session(job_id: str) -> bool:
    safe_job_id = _safe_job_id(job_id)
    folder = get_review_session_folder(safe_job_id)

    meta = _read_json(os.path.join(folder, "metadata.json"), {})
    records = _read_json(os.path.join(folder, "current_review.json"), [])
    audit_records = _read_json(os.path.join(folder, "audit_log.json"), [])
    control_records = _read_json(os.path.join(folder, "extraction_control.json"), [])

    df = _records_to_df(records)
    if not isinstance(df, pd.DataFrame) or df.empty:
        st.error("No saved review data found for this Job ID.")
        return False

    site = meta.get("site", st.session_state.get("site", SITES[0]))
    if site not in SITES:
        site = SITES[0]

    st.session_state.job_id = safe_job_id
    st.session_state.site = site
    st.session_state.product_name = meta.get("product_name", "")
    st.session_state.batch_record_name = meta.get("batch_record_name", "")
    st.session_state.batch_record_fp = meta.get("batch_record_fp", None)
    st.session_state.reviewer_name = meta.get("reviewer_name", "")
    st.session_state.review_status = meta.get("review_status", "Draft")
    st.session_state.reviewed_at = meta.get("reviewed_at", "")
    st.session_state.review_comment = meta.get("review_comment", "")
    st.session_state.validation_status = meta.get("validation_status", "Not validated yet")
    st.session_state.baseline_hash = meta.get("baseline_hash", "")
    st.session_state.extraction_source = meta.get("extraction_source", "saved_review")
    st.session_state.extraction_started_at = meta.get("extraction_started_at", "")
    st.session_state.extraction_ended_at = meta.get("extraction_ended_at", "")
    st.session_state.extraction_seconds = meta.get("extraction_seconds", None)
    st.session_state.extraction_rows = int(meta.get("extraction_rows", len(df)) or len(df))
    st.session_state.last_saved_at = meta.get("last_saved_at", "")

    df = ensure_workflow_columns(df)
    df = ensure_internal_row_ids(df)
    st.session_state.df = df

    control_df = _records_to_df(control_records)
    st.session_state.extraction_control_df = control_df if not control_df.empty else build_extraction_control_df(df)

    st.session_state.audit_log = audit_records if isinstance(audit_records, list) else []
    st.session_state.active_schema = get_active_schema(site)
    st.session_state.active_columns = derive_columns_from_schema_or_df(st.session_state.df, st.session_state.active_schema)
    st.session_state.validation = validate_df(st.session_state.df, st.session_state.active_schema)
    st.session_state.user_validated = False
    st.session_state.view_step = "Review"
    set_blocking_messages([])

    audit("saved_review_resumed", f"job_id={safe_job_id}; rows={len(df)}")
    autosave_review_session("saved_review_resumed")
    return True


# =============================================================================
# Top bar
# =============================================================================
if not st.session_state.get("full_screen_review", False):
    top_left, top_center, top_right = st.columns([1, 8, 1])

    with top_left:
        st.markdown('<div class="kebab-btn">', unsafe_allow_html=True)
        if st.button("⋮", help="Menu"):
            st.session_state.show_menu = not st.session_state.show_menu
        st.markdown("</div>", unsafe_allow_html=True)

    with top_center:
        st.markdown(
            """
            <div style="text-align:center; margin-top: 6px;">
              <div style="
                font-size: 46px;
                font-weight: 1000;
                letter-spacing: -1px;
                color: #ffffff;
                text-shadow: 0 3px 10px rgba(0,0,0,0.55);
                text-transform: uppercase;
              ">
                Auto BR Extractor Tool
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with top_right:
        st.write("")

if (not st.session_state.get("full_screen_review", False)) and st.session_state.show_menu:
    with st.expander("Menu", expanded=True):

        st.markdown("### Simplification support")
        st.write(
            """
This tool is designed to support the **Batch Record Simplification** exercise by extracting structured data
from **empty Batch Record templates**, with the output aligned to the toolbox template structure.
            """
        )

        st.markdown("---")

        st.markdown("### About this tool")
        st.write(
            """
- Start from Batch Record (PDF)
- Review, correct, add rows and remove rows inside the app only
- Every saved modification is recorded in the audit log
- Validation status only changes when you click "Validate table" or "Approve & lock"
- Approve + Export final Excel
            """
        )

        st.markdown("---")
        st.button("🔄 Reset session", on_click=reset_all)
        st.write()

# =============================================================================
# Steps nav (ALWAYS ENABLED)
# =============================================================================
if not st.session_state.get("full_screen_review", False):
    b1, b2, b3, b4, b5 = st.columns([2, 1, 1, 1, 2])
    with b2:
        go_setup = st.button("Step 1 · Setup")
    with b3:
        go_review = st.button("Step 2 · Review")
    with b4:
        go_export = st.button("Step 3 · Export Excel")

    if go_setup:
        set_blocking_messages([])
        st.session_state.view_step = "Setup"
    elif go_review:
        go_or_warn("Review")
    elif go_export:
        go_or_warn("Export Excel")

# Full-screen mode intentionally hides the global navigation.
# The review workspace contains a clear "Exit full screen" button.
enforce_step_guard_soft()


# =============================================================================
# Summary cards
# =============================================================================
def render_summary_cards():
    df = st.session_state.df
    schema = st.session_state.active_schema or {}

    rows = f"{len(df):,}" if isinstance(df, pd.DataFrame) else "-"
    skipped = st.session_state.skipped_pages or []
    skipped_txt = ", ".join(map(str, skipped)) if skipped else "-"

    page_count_txt = "-"
    page_col = (schema.get("page_col") or "").strip()
    if isinstance(df, pd.DataFrame) and page_col and page_col in df.columns:
        pages = pd.to_numeric(df[page_col], errors="coerce").dropna()
        if len(pages) > 0:
            page_count_txt = str(int(pages.round().astype(int).nunique()))

    sec = st.session_state.extraction_seconds
    source = st.session_state.extraction_source or "-"
    timing_txt = "-"
    if sec is not None:
        timing_txt = f"{sec:.2f}s"
        if st.session_state.extraction_rows and sec > 0:
            timing_txt += f" · {(st.session_state.extraction_rows / sec):.2f} rows/s"

    k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
    with k1:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">File name</div><div class="metric-value">{st.session_state.batch_record_name or "-"}</div></div>',
            unsafe_allow_html=True,
        )
    with k2:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Date / user</div><div class="metric-value">{time.strftime("%Y-%m-%d")} · {st.session_state.reviewer_name or "-"}</div></div>',
            unsafe_allow_html=True,
        )
    with k3:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Total rows</div><div class="metric-value">{rows}</div></div>',
            unsafe_allow_html=True,
        )
    with k4:
        st.markdown(
            f'<div class="metric-card">'
            f'  <div class="metric-label">Product name</div>'
            f'  <div class="metric-value">{st.session_state.product_name or "-"}</div>'
            f'</div>',
            unsafe_allow_html=True
        )
    with k5:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Number of pages</div><div class="metric-value">{page_count_txt}</div></div>',
            unsafe_allow_html=True,
        )
    with k6:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Skipped pages</div><div class="metric-value">{skipped_txt}</div></div>',
            unsafe_allow_html=True,
        )
    with k7:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Extraction</div><div class="metric-value">{source} · {timing_txt}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="metric-card"><div class="metric-label">Validation</div>'
        f'<div class="metric-value">{st.session_state.validation_status} · Status: {st.session_state.review_status}</div></div>',
        unsafe_allow_html=True,
    )
    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)


# =============================================================================
# Step 2 helper: reviewer propagation
# =============================================================================
def propagate_reviewer_to_df(fill_all_rows: bool = False):
    df = st.session_state.df
    if not isinstance(df, pd.DataFrame):
        return

    schema = st.session_state.active_schema or {}
    ok_col = (schema.get("review_ok_col") or REVIEW_OK_COL).strip()
    reviewer_col = (schema.get("reviewer_col") or REVIEWER_COL).strip()
    reviewed_at_col = (schema.get("reviewed_at_col") or REVIEWED_AT_COL).strip()

    who = (st.session_state.reviewer_name or "").strip()
    if not who:
        return
    if reviewer_col not in df.columns:
        return

    if fill_all_rows:
        target_idx = df.index
    else:
        empty_mask = df[reviewer_col].isna() | (df[reviewer_col].astype(str).str.strip() == "")
        target_idx = df.index[empty_mask]

    if len(target_idx) > 0:
        df.loc[target_idx, reviewer_col] = who

    if ok_col in df.columns and reviewed_at_col in df.columns:
        ok_mask = coerce_review_ok_series(df[ok_col])
        missing_dt = df[reviewed_at_col].isna() | (df[reviewed_at_col].astype(str).str.strip() == "")
        dt_idx = df.index[ok_mask & missing_dt]
        if len(dt_idx) > 0:
            df.loc[dt_idx, reviewed_at_col] = time.strftime("%Y-%m-%d %H:%M:%S")

    st.session_state.df = df


# =============================================================================
# Step 1: Setup
# =============================================================================
def render_setup_step():
    st.markdown('<div class="panel">', unsafe_allow_html=True)
    step_title("Setup", "Step 1: New extraction")

    if st.session_state.blocking_msgs:
        st.warning("You can’t continue yet:\n\n" + "\n".join(st.session_state.blocking_msgs))

    st.markdown("#### New extraction (PDF)")

    # Important notice must be visible before the reviewer enters product name
    # and before any upload field appears.
    st.markdown(
        """
        <div style="
            background: rgba(10, 18, 35, 0.78);
            border: 1px solid rgba(255,255,255,0.32);
            border-radius: 16px;
            padding: 18px;
            box-shadow: 0 10px 28px rgba(0,0,0,0.35);
            backdrop-filter: blur(10px);
            margin: 8px 0 22px 0;
            color: #ffffff !important;
        ">
            <div style="font-weight:1000; font-size:18px; margin-bottom:10px; color:#ffffff !important;">
                Important notice
            </div>
            <ul style="margin:0; padding-left:20px; line-height:1.6; color:#ffffff !important; font-weight:850;">
                <li style="color:#ffffff !important;">This tool works only with <b>EMPTY Batch Record templates</b>.</li>
                <li style="color:#ffffff !important;">Do <b>NOT</b> upload filled, handwritten, completed, signed, or executed Batch Records.</li>
                <li style="color:#ffffff !important;">Scanned PDFs may impact extraction quality and structure.</li>
                <li style="color:#ffffff !important;">Please upload the original digital empty template in PDF format.</li>
            </ul>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns([1.2, 1.8])
    with c1:
        new_site = st.selectbox("Select site", SITES, index=SITES.index(st.session_state.site))
    with c2:
        new_product = st.text_input(
            "Product name (required before upload)",
            value=st.session_state.product_name,
            placeholder="Type product name first…",
        )

    if new_product != st.session_state.product_name:
        st.session_state.product_name = new_product
        if isinstance(st.session_state.df, pd.DataFrame):
            st.session_state.df = apply_ui_metadata_to_df(
                st.session_state.site,
                st.session_state.df,
                st.session_state.batch_record_name or "record",
            )
            audit("product_name_changed", f"product={new_product}")
            autosave_review_session("product_name_changed")

    if new_site != st.session_state.site:
        st.session_state.site = new_site
        st.session_state.active_schema = get_active_schema(new_site)
        st.session_state.active_columns = derive_columns_from_schema_or_df(st.session_state.df, st.session_state.active_schema)
        st.session_state.batch_record_fp = None

        if isinstance(st.session_state.df, pd.DataFrame):
            st.session_state.df = apply_ui_metadata_to_df(
                st.session_state.site,
                st.session_state.df,
                st.session_state.batch_record_name or "record",
            )
            audit("site_changed", f"site={new_site}")
            autosave_review_session("site_changed")

    if needs_product():
        st.markdown(
            """
            <div style="
                background: rgba(10, 18, 35, 0.72);
                border: 1px solid rgba(255,255,255,0.25);
                border-radius: 12px;
                padding: 14px 18px;
                color: #ffffff !important;
                font-weight: 900;
                font-size: 16px;
                box-shadow: 0 8px 22px rgba(0,0,0,0.25);
                margin-top: 18px;
            ">
                Please enter the Product name first. The Batch Record upload field will appear after the product name is filled.
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown("#### Upload Batch Record (single file)")
        uploaded = st.file_uploader(
            label="Upload Batch Record (PDF)",
            type=["pdf"],
            key="uploader_batch_record",
            help="Extraction starts automatically after upload.",
        )
        if uploaded is not None:
            st.session_state.batch_record_bytes = uploaded.read()
            st.session_state.batch_record_name = uploaded.name
            audit("batch_record_uploaded", uploaded.name)
            with st.spinner("Auto-extracting..."):
                time.sleep(0.05)
                maybe_auto_extract()
            autosave_review_session("new_extraction")
            st.success(f"Loaded, extracted and autosaved ✅ Job ID: {st.session_state.job_id}")
            set_blocking_messages([])

    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
    next_disabled = needs_product() or needs_data()

    col_next = st.columns([1])[0]
    with col_next:
        if st.button("➡ Next: Go to Step 2 · Review", disabled=next_disabled, use_container_width=True):
            set_blocking_messages([])
            st.session_state.view_step = "Review"
            st.rerun()

        if next_disabled:
            missing = []
            if needs_product():
                missing.append("Product name")
            if needs_data():
                missing.append("Upload Batch Record (PDF)")
            st.caption("Complete required fields: " + ", ".join(missing))

    st.markdown("</div>", unsafe_allow_html=True)

# =============================================================================
# Step 2: Review
# =============================================================================
def render_review_metadata_panel():
    st.subheader("Reviewer / status")

    cA, cB, cC = st.columns([1.3, 1.0, 1.0])

    with cA:
        new_reviewer = st.text_input(
            "Reviewer name",
            value=st.session_state.reviewer_name,
            placeholder="Required for approval/export/download",
        )

    if new_reviewer != st.session_state.reviewer_name:
        st.session_state.reviewer_name = new_reviewer
        propagate_reviewer_to_df(fill_all_rows=False)
        mark_dirty("reviewer_name_changed")
        audit("reviewer_name_changed", f"reviewer={new_reviewer}")
        autosave_review_session("reviewer_name_changed")

    with cB:
        old_status = st.session_state.review_status
        new_status = st.selectbox(
            "Review status",
            REVIEW_STATUS_CHOICES,
            index=REVIEW_STATUS_CHOICES.index(st.session_state.review_status),
        )
        if new_status != old_status:
            st.session_state.review_status = new_status
            audit("review_status_changed", f"old={old_status}; new={new_status}; reviewer={st.session_state.reviewer_name}")
            autosave_review_session("review_status_changed")

    with cC:
        st.caption(f"Reviewed at: {st.session_state.reviewed_at or '-'}")
        if st.session_state.last_saved_at:
            st.caption(f"Last autosave: {st.session_state.last_saved_at}")

    old_comment = st.session_state.review_comment
    new_comment = st.text_area(
        "Reviewer global comment",
        value=st.session_state.review_comment,
        placeholder="Optional comment...",
        height=80,
    )
    if new_comment != old_comment:
        st.session_state.review_comment = new_comment
        audit("review_comment_changed", f"reviewer={st.session_state.reviewer_name}")
        autosave_review_session("review_comment_changed")

    c1, c2, c3 = st.columns([1.1, 1.1, 2.0])

    with c1:
        if st.button("✅ Approve & lock"):
            if needs_product():
                st.error("Product name is required to approve. Please fill it in Step 1.")
                return

            if not st.session_state.reviewer_name.strip():
                st.error("Reviewer name is required to approve.")
                return

            st.session_state.validation = validate_df(
                st.session_state.df,
                st.session_state.active_schema,
            )
            st.session_state.user_validated = True

            err_n = len(st.session_state.validation.errors)
            warn_n = len(st.session_state.validation.warnings)

            if err_n > 0:
                st.session_state.validation_status = f"Validated with notes ⚠️ ({err_n} items)"
                st.error("Approval blocked: some required fields are missing.")
                with st.expander("Show details"):
                    st.dataframe(
                        st.session_state.validation.errors,
                        use_container_width=True,
                        hide_index=True,
                    )
                audit("validated_on_approve_blocked", f"errors={err_n}; warnings={warn_n}")
                return

            st.session_state.validation_status = "Validated ✅"

            df = st.session_state.df
            schema = st.session_state.active_schema or {}
            ok_col = (schema.get("review_ok_col") or REVIEW_OK_COL).strip()

            not_reviewed = (
                int((~coerce_review_ok_series(df[ok_col])).sum())
                if df is not None and ok_col in df.columns
                else 0
            )

            if not_reviewed > 0:
                st.error(f"Approval blocked: {not_reviewed} rows are not reviewed yet.")
                return

            propagate_reviewer_to_df(fill_all_rows=False)

            st.session_state.review_status = "Approved"
            st.session_state.reviewed_at = time.strftime("%Y-%m-%d %H:%M:%S")
            st.session_state.baseline_hash = df_content_hash(st.session_state.df)

            audit(
                "approved",
                f"reviewer={st.session_state.reviewer_name}; hash={st.session_state.baseline_hash}",
            )
            autosave_review_session("approved")

            st.success("Approved and autosaved ✅")
            st.rerun()

    with c2:
        if st.button("🔓 Unlock (set Draft)"):
            st.session_state.review_status = "Draft"
            st.session_state.reviewed_at = ""
            audit("unlocked", "status set to Draft")
            autosave_review_session("unlocked")
            st.rerun()

    with c3:
        st.caption(
            "Tip: Check 'Reviewed & OK' when a row is verified. On Apply / Save, Reviewer and Reviewed at are automatically filled."
        )


def build_empty_review_row(cols: List[str]) -> Dict[str, Any]:
    """Create one empty business row for controlled manual addition inside the app."""
    row: Dict[str, Any] = {}
    for col in cols:
        if col == INTERNAL_ROW_ID_COL:
            row[col] = str(uuid.uuid4())[:8]
        elif col == REVIEW_OK_COL:
            row[col] = False
        elif col == REVIEWER_COL:
            row[col] = ""
        elif col == REVIEWED_AT_COL:
            row[col] = ""
        elif col == REVIEWER_NOTE_COL:
            row[col] = "Manual row added during in-app review"
        elif col == "PDF name":
            row[col] = st.session_state.batch_record_name or ""
        elif col == "Process step" and st.session_state.site in {"Arklow", "Bolbec", "Toledo"}:
            row[col] = st.session_state.product_name or ""
        else:
            row[col] = ""
    return row


def audit_cell_changes(
    before: pd.DataFrame,
    after: pd.DataFrame,
    editable_cols: List[str],
    reviewer: str,
):
    """Log old/new values for every edited cell shown in the current page."""
    if before.empty or after.empty:
        return

    for idx in after.index:
        if idx not in before.index:
            continue

        row_id = str(after.at[idx, INTERNAL_ROW_ID_COL]) if INTERNAL_ROW_ID_COL in after.columns else str(idx)

        for col in editable_cols:
            if col not in before.columns or col not in after.columns:
                continue

            old_val = "" if pd.isna(before.at[idx, col]) else str(before.at[idx, col])
            new_val = "" if pd.isna(after.at[idx, col]) else str(after.at[idx, col])

            if old_val != new_val:
                audit(
                    "cell_changed",
                    f"row_id={row_id}; row_index={idx}; column={col}; old={old_val}; new={new_val}; reviewer={reviewer}",
                )


def render_in_app_row_management_panel(is_locked: bool, cols: List[str]):
    """Controlled add/remove actions. No offline Excel review."""
    st.subheader("Controlled row management")

    if is_locked:
        st.info("Rows cannot be added or removed after approval. Click 'Unlock (set Draft)' if a correction is needed.")
        return

    if not (st.session_state.reviewer_name or "").strip():
        st.warning("Enter reviewer name before adding or removing rows.")
        return

    tab_add, tab_remove_help = st.tabs(["➕ Add row", "🗑️ Remove rows"])

    with tab_add:
        st.caption("Add a row only when the reviewer identifies a missing item in the extracted pre-draft.")

        new_row_cols = [c for c in cols if c not in {INTERNAL_ROW_ID_COL, REVIEWER_COL, REVIEWED_AT_COL}]
        new_row = build_empty_review_row(new_row_cols)
        new_row_df = pd.DataFrame([new_row])

        if REVIEW_OK_COL in new_row_df.columns:
            new_row_df[REVIEW_OK_COL] = coerce_review_ok_series(new_row_df[REVIEW_OK_COL])

        edited_new_row = st.data_editor(
            new_row_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            key="add_row_editor",
            column_config={
                REVIEW_OK_COL: st.column_config.CheckboxColumn(
                    "Reviewed & OK",
                    help="Keep unchecked until the added row is verified.",
                )
            } if REVIEW_OK_COL in new_row_df.columns else None,
        )

        c_add_1, c_add_2 = st.columns([1, 3])
        with c_add_1:
            if st.button("➕ Add row to table", use_container_width=True):
                df = ensure_internal_row_ids(st.session_state.df)
                row_to_add = edited_new_row.iloc[0].to_dict()
                row_to_add[INTERNAL_ROW_ID_COL] = str(uuid.uuid4())[:8]

                if REVIEWER_COL in df.columns:
                    row_to_add[REVIEWER_COL] = st.session_state.reviewer_name.strip()
                if REVIEWED_AT_COL in df.columns:
                    row_to_add[REVIEWED_AT_COL] = ""
                if REVIEW_OK_COL in df.columns:
                    row_to_add[REVIEW_OK_COL] = False

                # Ensure all existing columns are present and ordered.
                for col in df.columns:
                    row_to_add.setdefault(col, "")

                new_df = pd.concat([df, pd.DataFrame([row_to_add])[df.columns]], ignore_index=True)
                st.session_state.df = ensure_internal_row_ids(new_df)
                st.session_state.extraction_rows = int(len(st.session_state.df))
                mark_dirty("row_added")
                audit(
                    "row_added",
                    f"row_id={row_to_add[INTERNAL_ROW_ID_COL]}; reviewer={st.session_state.reviewer_name}; total_rows={len(st.session_state.df)}",
                )
                autosave_review_session("row_added")
                st.success("Row added and autosaved ✅")
                st.rerun()

        with c_add_2:
            st.caption("The added row is kept as not reviewed until it is checked and approved in the table.")

    with tab_remove_help:
        st.caption("To remove rows, tick the temporary 'Remove row' checkbox in the review table below, then click the remove button.")



def render_review_step():
    st.markdown('<div class="panel">', unsafe_allow_html=True)
    step_title("Review", "Step 2: Controlled Excel-like review")

    msgs = []
    if needs_product():
        msgs.append("• Please enter **Product name** first (Step 1).")
    if needs_data():
        msgs.append("• Please **upload Batch Record (PDF)** first (Step 1).")
    if msgs:
        st.warning("You can’t continue yet:\n\n" + "\n".join(msgs))
        st.markdown("</div>", unsafe_allow_html=True)
        return

    st.session_state.df = ensure_internal_row_ids(st.session_state.df)

    # Stable mode: no Streamlit components and no browser full-screen.
    # We only enlarge the editable table height inside the normal app page.
    st.session_state.full_screen_review = False
    full_screen = False
    if full_screen:
        inject_review_fullscreen_css()
        st.markdown(
            """
            <div class="fullscreen-review-banner">
              <div style="font-weight:1000; font-size:22px; color:white;">🖥️ Full-screen controlled review mode</div>
              <div style="font-weight:800; opacity:0.92; color:white;">
                Use the large table workspace for comfortable correction. All changes remain controlled, audited and autosaved.
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        render_fullscreen_instruction()
        st.markdown(
            '<div class="fullscreen-toolbar-note">Tip: for real browser full-screen, press F11. This version avoids custom JS components to prevent component errors.</div>',
            unsafe_allow_html=True,
        )

        fs1, fs2, fs3, fs4, fs5 = st.columns([1.35, 0.85, 1.45, 1.05, 0.85])
        with fs1:
            new_reviewer_fs = st.text_input(
                "Reviewer name",
                value=st.session_state.reviewer_name,
                placeholder="Required before saving",
                key="reviewer_name_fullscreen",
            )
            if new_reviewer_fs != st.session_state.reviewer_name:
                st.session_state.reviewer_name = new_reviewer_fs
                propagate_reviewer_to_df(fill_all_rows=False)
                mark_dirty("reviewer_name_changed_fullscreen")
                audit("reviewer_name_changed", f"reviewer={new_reviewer_fs}; source=fullscreen")
                autosave_review_session("reviewer_name_changed_fullscreen")
        with fs2:
            st.metric("Status", st.session_state.review_status)
        with fs3:
            st.caption(f"Job ID: {st.session_state.job_id} · Last autosave: {st.session_state.last_saved_at or '-'}")
            st.caption(f"File: {st.session_state.batch_record_name or '-'}")
        with fs4:
            st.session_state.review_table_height_fullscreen = st.number_input(
                "Table height",
                min_value=650,
                max_value=1400,
                value=int(st.session_state.review_table_height_fullscreen),
                step=50,
                help="Adjust the grid height for your monitor.",
            )
        with fs5:
            if st.button("↩ Exit app full screen", use_container_width=True):
                st.session_state.full_screen_review = False
                st.rerun()
    else:
        render_summary_cards()
        render_review_metadata_panel()
        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
        render_extraction_control_pdf_download("step2")
        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

    df = ensure_internal_row_ids(st.session_state.df)
    schema = st.session_state.active_schema or {}
    business_cols = get_schema_columns_for_export(df, schema) or [c for c in df.columns if c not in INTERNAL_COLS]

    search_cols = [c for c in derive_search_cols(schema, business_cols) if c in df.columns]
    if not search_cols:
        search_cols = business_cols

    ok_col = (schema.get("review_ok_col") or REVIEW_OK_COL).strip()
    reviewer_col = (schema.get("reviewer_col") or REVIEWER_COL).strip()
    reviewed_at_col = (schema.get("reviewed_at_col") or REVIEWED_AT_COL).strip()
    note_col = (schema.get("review_note_col") or REVIEWER_NOTE_COL).strip()

    if ok_col in df.columns:
        df[ok_col] = coerce_review_ok_series(df[ok_col])
        st.session_state.df = df

    is_locked = (st.session_state.review_status == "Approved")

    st.subheader("Excel-like review workspace · enlarged table")
    st.caption(
        "Edit cells directly in the controlled table. Enlarge the table with the Table size option. Use the toolbar to insert rows above/below, "
        "add a row at the end, or delete the selected row. No offline Excel review is required."
    )

    # Reviewer comfort: enlarge only the table area, no custom/full-screen component.
    size_col1, size_col2 = st.columns([1.4, 3.8])
    with size_col1:
        size_options = ["Comfort", "Large", "Extra large"]
        current_size = st.session_state.get("review_table_size_mode", "Large")
        size_index = size_options.index(current_size) if current_size in size_options else 1
        st.session_state.review_table_size_mode = st.selectbox(
            "Table size",
            size_options,
            index=size_index,
            help="Use Large or Extra large to make the review table easier to edit. No browser full-screen or custom component is used.",
        )
        height_by_mode = {"Comfort": 650, "Large": 850, "Extra large": 1100}
        st.session_state.review_table_height_normal = height_by_mode[st.session_state.review_table_size_mode]
    with size_col2:
        st.caption(
            "Use the table size option to enlarge the review workspace safely. "
            "For maximum screen space, the reviewer can also use the browser zoom/F11 manually if allowed."
        )

    if not is_locked and not (st.session_state.reviewer_name or "").strip():
        st.warning("Enter reviewer name before saving modifications, inserting rows, adding rows, or deleting rows.")

    # -------------------------------------------------------------------------
    # Filters
    # -------------------------------------------------------------------------
    f1, f2, f3, f4 = st.columns([1.4, 1.8, 1.0, 1.4])
    with f1:
        view_options = [
            "All rows",
            "Not reviewed only",
            "Reviewed only",
            "Rows with validation issues",
            "Rows with reviewer notes",
        ]
        default_idx = view_options.index(st.session_state.view_mode) if st.session_state.view_mode in view_options else 1
        st.session_state.view_mode = st.selectbox(
            "View mode",
            view_options,
            index=default_idx,
            disabled=is_locked,
        )
    with f2:
        st.session_state.search = st.text_input(
            "Search",
            value=st.session_state.search,
            placeholder="Search by data title, tag, process step, page…",
            disabled=is_locked,
        )
    with f3:
        page_size_options = [25, 50, 100, 200, 500]
        default_page_size = 200 if st.session_state.get("review_table_size_mode") in {"Large", "Extra large"} else 100
        default_page_idx = page_size_options.index(default_page_size)
        page_size = st.selectbox("Rows/page", page_size_options, index=default_page_idx, disabled=is_locked)
    with f4:
        st.caption("Toolbar actions are audited and autosaved server-side after you click Apply / Save.")

    view = df.copy()

    if st.session_state.view_mode == "Not reviewed only" and ok_col in view.columns:
        view[ok_col] = coerce_review_ok_series(view[ok_col])
        view = view.loc[~view[ok_col]]
    elif st.session_state.view_mode == "Reviewed only" and ok_col in view.columns:
        view[ok_col] = coerce_review_ok_series(view[ok_col])
        view = view.loc[view[ok_col]]
    elif st.session_state.view_mode == "Rows with validation issues":
        validation = validate_df(drop_internal_cols(df), schema)
        flagged = set(validation.flagged_row_ids)
        view = view.loc[[i for i in view.index if int(i) in flagged]]
    elif st.session_state.view_mode == "Rows with reviewer notes" and note_col in view.columns:
        view = view.loc[view[note_col].fillna("").astype(str).str.strip() != ""]

    q = (st.session_state.search or "").strip().lower()
    if q:
        mask = pd.Series(False, index=view.index)
        for c in search_cols:
            if c in view.columns:
                mask = mask | view[c].astype(str).str.lower().str.contains(q, na=False)
        view = view[mask]

    display_business_cols = [c for c in business_cols if c in view.columns]
    view = view[display_business_cols + [INTERNAL_ROW_ID_COL]].copy()
    if ok_col in view.columns:
        view[ok_col] = coerce_review_ok_series(view[ok_col])

    total = len(view)
    max_pages = max(1, int((total + page_size - 1) / page_size))
    page_number = st.number_input(
        "Page",
        min_value=1,
        max_value=max_pages,
        value=min(int(st.session_state.get("review_page_number", 1)), max_pages),
        step=1,
        disabled=is_locked,
    )
    st.session_state.review_page_number = int(page_number)

    start = (int(page_number) - 1) * page_size
    end = start + page_size
    page_view = view.iloc[start:end].copy()

    total_not_reviewed = int((~coerce_review_ok_series(df[ok_col])).sum()) if ok_col in df.columns else 0
    st.caption(
        f"Filtered rows: {total:,} · Showing {len(page_view):,} rows · Page {int(page_number)}/{max_pages} · "
        f"Not reviewed remaining: {total_not_reviewed:,} · Last autosave: {st.session_state.last_saved_at or '-'}"
    )

    if is_locked:
        st.info("Locked because status is Approved. Click 'Unlock (set Draft)' to edit.")

    # -------------------------------------------------------------------------
    # Excel-like toolbar: no ugly Insert/Delete columns inside the table.
    # The reviewer uses the visible row number to select the target row.
    # -------------------------------------------------------------------------
    DISPLAY_ROW_COL = "__display_row_number__"
    page_view_for_editor = page_view.copy()

    # Keep only a neutral Row # column. No colored row band.
    row_numbers = list(range(start + 1, start + 1 + len(page_view_for_editor)))
    page_view_for_editor.insert(0, DISPLAY_ROW_COL, row_numbers)

    has_page_rows = len(page_view_for_editor) > 0
    min_row_number = int(start + 1) if has_page_rows else 1
    max_row_number = int(start + len(page_view_for_editor)) if has_page_rows else 1
    previous_target = int(st.session_state.get("toolbar_target_row_number", min_row_number))
    previous_target = min(max(previous_target, min_row_number), max_row_number)

    st.markdown("#### Row toolbar")
    tb1, tb2, tb3, tb4, tb5, tb6 = st.columns([1.15, 1.1, 1.1, 1.1, 1.1, 1.4])
    with tb1:
        target_row_number = st.number_input(
            "Target row #",
            min_value=min_row_number,
            max_value=max_row_number,
            value=previous_target,
            step=1,
            disabled=is_locked or not has_page_rows,
            help="Use the Row # shown in the first column of the table below.",
        )
        st.session_state.toolbar_target_row_number = int(target_row_number)
    with tb2:
        insert_above_clicked = st.button("➕ Insert above", disabled=is_locked or not has_page_rows, use_container_width=True)
    with tb3:
        insert_below_clicked = st.button("➕ Insert below", disabled=is_locked or not has_page_rows, use_container_width=True)
    with tb4:
        delete_clicked = st.button("🗑️ Delete row", disabled=is_locked or not has_page_rows, use_container_width=True)
    with tb5:
        add_end_clicked = st.button("➕ Add at end", disabled=is_locked, use_container_width=True)
    with tb6:
        save_clicked = st.button("💾 Apply / Save", disabled=is_locked, use_container_width=True)

    st.caption(
        "How to use: choose the Target row # from the first column, then click Insert above, Insert below, or Delete row. "
        "Cell edits on the current page are saved at the same time."
    )

    selected_row_id = None
    if has_page_rows:
        selected_pos_on_page = int(target_row_number) - (start + 1)
        if 0 <= selected_pos_on_page < len(page_view):
            selected_row_id = str(page_view.iloc[selected_pos_on_page][INTERNAL_ROW_ID_COL])

    # -------------------------------------------------------------------------
    # Editable table
    # -------------------------------------------------------------------------
    disabled_cols = [DISPLAY_ROW_COL, INTERNAL_ROW_ID_COL, reviewer_col, reviewed_at_col]
    # PDF name is source metadata; keep it protected in the controlled review table.
    if "PDF name" in page_view_for_editor.columns:
        disabled_cols.append("PDF name")
    if is_locked:
        disabled_cols = list(page_view_for_editor.columns)

    inject_review_grid_readability_css()

    grid_height = int(st.session_state.review_table_height_normal)

    if HAS_AGGRID:
        # AG Grid gives clearer row segmentation than Streamlit's native data_editor:
        # visible grid lines, zebra rows, stronger header, clearer focus cell, and pinned row numbers.
        ag_grid_df = page_view_for_editor.copy()
        if ok_col in ag_grid_df.columns:
            ag_grid_df[ok_col] = coerce_review_ok_series(ag_grid_df[ok_col])

        gb = GridOptionsBuilder.from_dataframe(ag_grid_df)
        gb.configure_default_column(
            editable=not is_locked,
            filter=False,
            sortable=False,
            resizable=True,
            wrapText=True,
            autoHeight=True,
            minWidth=130,
        )

        for col in ag_grid_df.columns:
            editable = (not is_locked) and col not in disabled_cols

            if col == DISPLAY_ROW_COL:
                gb.configure_column(
                    col,
                    header_name="Row #",
                    editable=False,
                    pinned="left",
                    width=85,
                    minWidth=75,
                    maxWidth=95,
                    cellStyle={
                        "fontWeight": "900",
                        "backgroundColor": "#eef2ff",
                        "borderRight": "2px solid #94a3b8",
                        "textAlign": "center",
                    },
                )
            elif col == INTERNAL_ROW_ID_COL:
                # Keep traceability in data returned to Python, but hide it from reviewer view.
                gb.configure_column(col, hide=True, editable=False)
            elif col == ok_col:
                gb.configure_column(
                    col,
                    header_name="Reviewed & OK",
                    editable=editable,
                    pinned="left",
                    width=145,
                    minWidth=135,
                    cellRenderer=get_aggrid_review_ok_checkbox_renderer(),
                    cellStyle={
                        "fontWeight": "900",
                        "backgroundColor": "#f8fafc",
                        "borderRight": "2px solid #cbd5e1",
                        "justifyContent": "center",
                        "textAlign": "center",
                    },
                )
            elif col in {reviewer_col, reviewed_at_col}:
                gb.configure_column(col, editable=False, width=165, minWidth=150)
            elif col == "PDF name":
                gb.configure_column(col, editable=False, width=180, minWidth=160)
            elif col in {"Data title", "Data Title"}:
                gb.configure_column(col, editable=editable, width=320, minWidth=240)
            elif col in {"Process step", "Sub process step"}:
                gb.configure_column(col, editable=editable, width=220, minWidth=180)
            elif col in {note_col, REVIEWER_NOTE_COL}:
                gb.configure_column(col, editable=editable, width=280, minWidth=220)
            else:
                gb.configure_column(col, editable=editable, width=150, minWidth=110)

        gb.configure_grid_options(
            rowHeight=44,
            headerHeight=46,
            suppressRowClickSelection=True,
            ensureDomOrder=True,
            enableCellTextSelection=True,
            stopEditingWhenCellsLoseFocus=True,
        )

        grid_options = gb.build()

        st.markdown('<div class="review-grid-shell">', unsafe_allow_html=True)
        grid_response = AgGrid(
            ag_grid_df,
            gridOptions=grid_options,
            update_mode=GridUpdateMode.MODEL_CHANGED,
            data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
            fit_columns_on_grid_load=False,
            height=grid_height,
            allow_unsafe_jscode=True,
            theme="alpine",
            custom_css=get_aggrid_custom_css(),
            key=f"review_aggrid_toolbar_{st.session_state.review_page_number}_{st.session_state.view_mode}_{st.session_state.search}_{is_locked}",
        )
        st.markdown('</div>', unsafe_allow_html=True)

        edited_page = pd.DataFrame(grid_response.get("data", ag_grid_df))

        # Make sure hidden/internal columns and column order are preserved before saving.
        for col in page_view_for_editor.columns:
            if col not in edited_page.columns:
                edited_page[col] = page_view_for_editor[col].values
        edited_page = edited_page[page_view_for_editor.columns]
        if ok_col in edited_page.columns:
            edited_page[ok_col] = coerce_review_ok_series(edited_page[ok_col])

    else:
        st.caption("Stable native editor mode: dark table border added for clearer separation without colored row markers.")
        column_config = {
            DISPLAY_ROW_COL: st.column_config.NumberColumn(
                "Row #",
                help="Use this number in the toolbar above.",
                disabled=True,
                width="small",
            ),
            INTERNAL_ROW_ID_COL: None,
        }
        if ok_col in page_view_for_editor.columns:
            column_config[ok_col] = st.column_config.CheckboxColumn(
                "Reviewed & OK",
                help="Check only after the row has been verified against the PDF/source.",
            )

        edited_page = st.data_editor(
            page_view_for_editor,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            height=grid_height,
            disabled=[c for c in disabled_cols if c in page_view_for_editor.columns],
            column_config=column_config,
            key=f"review_editor_toolbar_{st.session_state.review_page_number}_{st.session_state.view_mode}_{st.session_state.search}_{is_locked}",
        )


    def _norm(v: Any) -> str:
        if pd.isna(v):
            return ""
        return str(v)

    def _make_blank_row(base_df: pd.DataFrame, reviewer: str, note: str) -> Dict[str, Any]:
        """Create a controlled blank row with the same schema as the current review table."""
        new_row: Dict[str, Any] = {}
        for col in base_df.columns:
            if col == INTERNAL_ROW_ID_COL:
                new_row[col] = str(uuid.uuid4())[:8]
            elif col == ok_col:
                new_row[col] = False
            elif col == reviewer_col:
                new_row[col] = reviewer
            elif col == reviewed_at_col:
                new_row[col] = ""
            elif col == note_col:
                new_row[col] = note
            elif col == "PDF name":
                new_row[col] = st.session_state.batch_record_name or ""
            elif col == "Process step" and st.session_state.site in {"Arklow", "Bolbec", "Toledo"}:
                new_row[col] = st.session_state.product_name or ""
            else:
                new_row[col] = ""
        return new_row

    def _apply_current_page_edits(df_work: pd.DataFrame, edited: pd.DataFrame, reviewer: str) -> Tuple[pd.DataFrame, int]:
        """Save edits made on the current page into the full dataframe and audit changed cells."""
        if not isinstance(edited, pd.DataFrame) or edited.empty:
            return df_work, 0

        edited = edited.drop(columns=[DISPLAY_ROW_COL], errors="ignore").copy()
        id_to_index = {str(row_id): int(idx) for idx, row_id in df_work[INTERNAL_ROW_ID_COL].items()}
        changed_count = 0
        now_ts = time.strftime("%Y-%m-%d %H:%M:%S")
        workflow_cols = {ok_col, reviewer_col, reviewed_at_col, note_col, INTERNAL_ROW_ID_COL}

        for _, edited_row in edited.iterrows():
            row_id = "" if pd.isna(edited_row.get(INTERNAL_ROW_ID_COL, "")) else str(edited_row.get(INTERNAL_ROW_ID_COL, "")).strip()
            if not row_id or row_id not in id_to_index:
                continue

            idx = id_to_index[row_id]
            business_changed = False
            old_ok = bool(coerce_review_ok_series(pd.Series([df_work.at[idx, ok_col]])).iloc[0]) if ok_col in df_work.columns else False
            new_ok = bool(coerce_review_ok_series(pd.Series([edited_row.get(ok_col, False)])).iloc[0]) if ok_col in edited_row.index else old_ok

            for col in business_cols:
                if col not in df_work.columns or col not in edited_row.index:
                    continue
                if col in {reviewer_col, reviewed_at_col}:
                    continue

                old_val = _norm(df_work.at[idx, col])
                new_val = _norm(edited_row[col])

                if old_val != new_val:
                    df_work.at[idx, col] = edited_row[col]
                    changed_count += 1
                    audit(
                        "cell_changed",
                        f"row_id={row_id}; row_index={idx}; column={col}; old={old_val}; new={new_val}; reviewer={reviewer}",
                    )
                    if col not in workflow_cols:
                        business_changed = True

            # Controlled reviewer/timestamp behavior for Reviewed & OK.
            # Quality rule:
            # - As soon as a row is checked as Reviewed & OK and saved, stamp Reviewer + Reviewed at.
            # - If a checked row is modified and remains checked, refresh Reviewed at to show the latest review time.
            # - If a row is unchecked, clear Reviewed at because the row is no longer reviewed.
            if ok_col in df_work.columns:
                previous_reviewed_at = ""
                if reviewed_at_col in df_work.columns and not pd.isna(df_work.at[idx, reviewed_at_col]):
                    previous_reviewed_at = str(df_work.at[idx, reviewed_at_col]).strip()

                if new_ok:
                    df_work.at[idx, ok_col] = True

                    if reviewer_col in df_work.columns:
                        df_work.at[idx, reviewer_col] = reviewer

                    should_stamp_time = (not old_ok) or business_changed or not previous_reviewed_at

                    if reviewed_at_col in df_work.columns and should_stamp_time:
                        df_work.at[idx, reviewed_at_col] = now_ts

                    if not old_ok:
                        audit(
                            "row_marked_reviewed",
                            f"row_id={row_id}; row_index={idx}; reviewer={reviewer}; reviewed_at={now_ts}",
                        )
                    elif business_changed:
                        audit(
                            "row_review_timestamp_refreshed",
                            f"row_id={row_id}; row_index={idx}; reviewer={reviewer}; reviewed_at={now_ts}; reason=business_cell_changed",
                        )

                else:
                    df_work.at[idx, ok_col] = False

                    if reviewed_at_col in df_work.columns and (old_ok or previous_reviewed_at):
                        df_work.at[idx, reviewed_at_col] = ""

                    # Keep Reviewer only if the row was edited during this save; otherwise clear it when unreviewed.
                    if reviewer_col in df_work.columns and not business_changed:
                        df_work.at[idx, reviewer_col] = ""

                    if old_ok:
                        audit(
                            "row_marked_unreviewed",
                            f"row_id={row_id}; row_index={idx}; reviewer={reviewer}; previous_reviewed_at={previous_reviewed_at}",
                        )
                    elif business_changed:
                        audit(
                            "row_kept_unreviewed_after_change",
                            f"row_id={row_id}; row_index={idx}; reviewer={reviewer}",
                        )

        return df_work, changed_count

    row_action_clicked = insert_above_clicked or insert_below_clicked or delete_clicked or add_end_clicked or save_clicked
    if not is_locked and row_action_clicked:
        reviewer = (st.session_state.reviewer_name or "").strip()
        if not reviewer:
            st.error("Reviewer name is required before saving or applying row actions.")
        else:
            try:
                df_work = ensure_internal_row_ids(st.session_state.df.copy()).reset_index(drop=True)
                df_work, changed_count = _apply_current_page_edits(df_work, edited_page, reviewer)

                inserted_count = 0
                added_count = 0
                deleted_count = 0

                if selected_row_id and (insert_above_clicked or insert_below_clicked or delete_clicked):
                    matching_indexes = df_work.index[df_work[INTERNAL_ROW_ID_COL].astype(str) == str(selected_row_id)].tolist()
                    if not matching_indexes:
                        st.warning("Selected row could not be found anymore. Refresh the page and try again.")
                    else:
                        selected_idx = int(matching_indexes[0])

                        if insert_above_clicked or insert_below_clicked:
                            direction = "above" if insert_above_clicked else "below"
                            insert_position = selected_idx if insert_above_clicked else selected_idx + 1
                            note = f"Manual row inserted {direction} during in-app review"
                            new_row = _make_blank_row(df_work, reviewer, note)
                            upper = df_work.iloc[:insert_position]
                            lower = df_work.iloc[insert_position:]
                            df_work = pd.concat([upper, pd.DataFrame([new_row])[df_work.columns], lower], ignore_index=True)
                            inserted_count = 1
                            audit(
                                "row_inserted",
                                f"direction={direction}; reference_row_id={selected_row_id}; new_row_id={new_row[INTERNAL_ROW_ID_COL]}; "
                                f"reviewer={reviewer}; total_rows={len(df_work)}",
                            )

                        if delete_clicked:
                            # Re-resolve the selected row after a possible insert; delete the reference row ID only.
                            delete_indexes = df_work.index[df_work[INTERNAL_ROW_ID_COL].astype(str) == str(selected_row_id)].tolist()
                            if delete_indexes:
                                delete_idx = int(delete_indexes[0])
                                title_col = "Data title" if "Data title" in df_work.columns else "Data Title" if "Data Title" in df_work.columns else ""
                                title_val = df_work.at[delete_idx, title_col] if title_col else ""
                                audit(
                                    "row_removed",
                                    f"row_id={selected_row_id}; row_index={delete_idx}; title={title_val}; reviewer={reviewer}",
                                )
                                df_work = df_work.drop(index=[delete_idx]).reset_index(drop=True)
                                deleted_count = 1

                if add_end_clicked:
                    new_row = _make_blank_row(df_work, reviewer, "Manual row added at end during in-app review")
                    df_work = pd.concat([df_work, pd.DataFrame([new_row])[df_work.columns]], ignore_index=True)
                    added_count = 1
                    audit(
                        "row_added",
                        f"row_id={new_row[INTERNAL_ROW_ID_COL]}; reviewer={reviewer}; source=toolbar_add_at_end; total_rows={len(df_work)}",
                    )

                st.session_state.df = ensure_internal_row_ids(df_work)
                st.session_state.extraction_rows = int(len(st.session_state.df))

                if inserted_count > 0:
                    mark_dirty("toolbar_row_inserted")
                elif added_count > 0:
                    mark_dirty("toolbar_row_added")
                elif deleted_count > 0:
                    mark_dirty("toolbar_row_deleted")
                elif changed_count > 0:
                    mark_dirty("toolbar_table_saved")

                audit(
                    "toolbar_saved_and_applied",
                    f"reviewer={reviewer}; inserted_rows={inserted_count}; added_rows={added_count}; "
                    f"deleted_rows={deleted_count}; changed_cells={changed_count}; target_row_number={st.session_state.toolbar_target_row_number}",
                )
                autosave_review_session("toolbar_saved_and_applied")
                st.success(
                    f"Saved and autosaved ✅ Inserted: {inserted_count}, Added at end: {added_count}, "
                    f"Deleted: {deleted_count}, Changed cells: {changed_count}"
                )
                st.rerun()

            except Exception as e:
                st.error(f"Could not save/apply toolbar action: {e}")
                st.code(traceback.format_exc())

    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

    if st.button("✅ Validate table", disabled=is_locked, use_container_width=True):
        validation_input = drop_internal_cols(st.session_state.df)
        st.session_state.validation = validate_df(validation_input, schema)
        st.session_state.baseline_hash = df_content_hash(validation_input)
        st.session_state.user_validated = True

        err_n = len(st.session_state.validation.errors)
        warn_n = len(st.session_state.validation.warnings)
        if err_n == 0:
            st.session_state.validation_status = "Table checks passed ✅"
            st.success(
                f"Table checks passed ✅ ({warn_n} warning(s)). "
                "The pre-draft Excel is available directly below."
            )
        else:
            st.session_state.validation_status = f"Table checks failed ❌ ({err_n} blocking items)"
            st.error(
                f"Table checks found {err_n} blocking item(s) and {warn_n} warning(s). "
                "Correct the blocking items before downloading the pre-draft Excel."
            )

        if err_n > 0:
            with st.expander("Show validation errors", expanded=True):
                st.dataframe(st.session_state.validation.errors, use_container_width=True, hide_index=True)
        if warn_n > 0:
            with st.expander("Show validation warnings", expanded=False):
                st.dataframe(st.session_state.validation.warnings, use_container_width=True, hide_index=True)

        audit("table_checks_run", f"errors={err_n}; warnings={warn_n}; hash={st.session_state.baseline_hash}")
        autosave_review_session("table_checks_run")

    render_working_predraft_excel_download()

    st.markdown("</div>", unsafe_allow_html=True)


# =============================================================================
# Step 3: Export (no freeze; cache bytes)
# =============================================================================
@st.cache_data(show_spinner=False)
def _build_excel_bytes_cached(
    df_csv: str,
    include_audit: bool,
    audit_csv: str,
    review_csv: str,
    control_csv: str,
    template_marker: str,
) -> Tuple[bytes, str, str]:
    """
    Returns (file_bytes, mime_type, extension).

    template_marker is included only to invalidate the cache when the
    template.xlsm file appears/disappears or is replaced — its value is
    derived from the template path + its mtime.
    """
    df = pd.read_csv(io.StringIO(df_csv))
    review_df = pd.read_csv(io.StringIO(review_csv)) if review_csv else pd.DataFrame()
    audit_df = pd.read_csv(io.StringIO(audit_csv)) if (include_audit and audit_csv) else None
    control_df = pd.read_csv(io.StringIO(control_csv)) if control_csv else build_extraction_control_df(df)
    return build_excel_bytes_macro(df, review_df, audit_df, include_audit, control_df_override=control_df)


def _template_cache_marker() -> str:
    """Cheap fingerprint of template.xlsm so the cached export refreshes if it changes."""
    if not _macro_template_available():
        return "no-template"
    try:
        st_info = os.stat(MACRO_TEMPLATE_PATH)
        return f"{MACRO_TEMPLATE_PATH}:{int(st_info.st_mtime)}:{st_info.st_size}"
    except Exception:
        return "template-stat-failed"



def make_working_predraft_review_sheet_df() -> pd.DataFrame:
    """Create metadata for a pre-draft Excel generated after table checks.

    Passing table checks confirms only that the configured structural controls
    succeeded. Business review and final approval remain pending in the app.
    """
    review_df = make_review_sheet_df().copy()

    df = st.session_state.df
    schema = st.session_state.active_schema or {}
    ok_col = (schema.get("review_ok_col") or REVIEW_OK_COL).strip()

    total_rows = int(len(df)) if isinstance(df, pd.DataFrame) else 0
    reviewed_rows = (
        int(coerce_review_ok_series(df[ok_col]).sum())
        if isinstance(df, pd.DataFrame) and ok_col in df.columns
        else 0
    )
    review_complete = total_rows > 0 and reviewed_rows == total_rows

    values = {
        "Status": "PRE-DRAFT - BUSINESS REVIEW PENDING",
        "Reviewed at": "",
        "Export type": "Pre-draft generated after successful table checks",
        "Technical table validation": "Passed",
        "Business review completed": "Yes" if review_complete else "No",
        "Rows reviewed": f"{reviewed_rows}/{total_rows}",
        "Final approval": "No",
        "Final use permitted": "No",
        "Official review location": "Controlled online application",
        "Export generated at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "Quality position": (
            "This file is a structured pre-draft generated after technical table checks. "
            "Business review and final approval remain controlled in the application."
        ),
    }

    for field, value in values.items():
        mask = review_df["Field"].astype(str).str.strip().eq(field)
        if mask.any():
            review_df.loc[mask, "Value"] = value
        else:
            review_df = pd.concat(
                [review_df, pd.DataFrame([{"Field": field, "Value": value}])],
                ignore_index=True,
            )

    return review_df


def add_working_predraft_notice_sheet(
    excel_bytes: bytes,
    excel_ext: str,
) -> bytes:
    """Add a clear pre-draft information sheet without presenting it as final."""
    source = io.BytesIO(excel_bytes)
    wb = openpyxl.load_workbook(source, keep_vba=(excel_ext == EXT_XLSM))

    notice_title = "PRE-DRAFT INFORMATION"
    legacy_titles = ["READ ME - NOT APPROVED", notice_title]
    for title in legacy_titles:
        if title in wb.sheetnames:
            wb.remove(wb[title])

    ws = wb.create_sheet(notice_title, 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F4E78"

    navy = PatternFill(fill_type="solid", fgColor="1F4E78")
    light_blue = PatternFill(fill_type="solid", fgColor="D9EAF7")
    amber = PatternFill(fill_type="solid", fgColor="FFF2CC")

    site = str(st.session_state.get("site", ""))
    product = str(st.session_state.get("product_name", ""))

    ws.merge_cells("A1:F2")
    ws["A1"] = f"PRE-DRAFT — {site} — {product}"
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A4:F5")
    ws["A4"] = (
        "This structured pre-draft was generated after successful technical table checks. "
        "Business review and final approval are not completed by this download and remain "
        "controlled in the application."
    )
    ws["A4"].font = Font(name="Arial", size=12, bold=True, color="1F1F1F")
    ws["A4"].fill = light_blue
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")

    rows = [
        ("Document type", "Structured pre-draft."),
        ("Purpose", "Consultation and preparation before completion of the controlled business review."),
        ("Technical table checks", "Passed at the time of export."),
        ("Business review", "Pending unless separately completed and approved in the application."),
        ("Final approved output", "Generated only after the controlled online review is completed and locked."),
        ("Site", site),
        ("Product", product),
        ("Source file", str(st.session_state.get("batch_record_name", ""))),
        ("Job ID", str(st.session_state.get("job_id", ""))),
        ("Reviewer", str(st.session_state.get("reviewer_name", ""))),
        ("Generated at", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("Tool version", TOOL_VERSION),
    ]

    start_row = 7
    for offset, (field, value) in enumerate(rows):
        row = start_row + offset
        ws.cell(row=row, column=1, value=field)
        ws.cell(row=row, column=2, value=value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

        ws.cell(row=row, column=1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=1).fill = amber
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=2).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20

    if "Extracted" in wb.sheetnames:
        wb["Extracted"].sheet_properties.tabColor = "5B9BD5"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def render_working_predraft_excel_download() -> None:
    """Render the pre-draft download immediately after successful table checks."""
    if st.session_state.review_status == "Approved":
        return

    # Nothing is displayed before the reviewer runs Validate table.
    if not st.session_state.get("user_validated", False):
        return

    validation = st.session_state.get("validation")
    blocking_errors = (
        len(validation.errors)
        if validation is not None and isinstance(getattr(validation, "errors", None), pd.DataFrame)
        else 0
    )
    if validation is None or blocking_errors > 0:
        return

    site = str(st.session_state.get("site", "Site")).strip() or "Site"
    product = str(st.session_state.get("product_name", "Product")).strip() or "Product"

    st.markdown(f"#### Pre-draft — {site} | {product}")
    st.caption(
        "Generated after successful table checks. Business review and final approval remain controlled in the application."
    )

    if not st.session_state.reviewer_name.strip():
        st.warning("Enter the reviewer name to download the pre-draft Excel.")
        return

    # Ensure application-generated review metadata is included in the integrity
    # check. If the table changed after validation, validation must be rerun.
    propagate_reviewer_to_df(fill_all_rows=False)

    current_hash = df_content_hash(drop_internal_cols(st.session_state.df))
    validated_hash = str(st.session_state.get("baseline_hash", ""))
    if not validated_hash or current_hash != validated_hash:
        st.session_state.user_validated = False
        st.session_state.validation_status = "Not validated yet"
        st.warning("The table changed after validation. Run **Validate table** again before downloading the pre-draft.")
        return

    df_csv = drop_internal_cols(st.session_state.df).to_csv(index=False)
    review_df = make_working_predraft_review_sheet_df()
    review_csv = review_df.to_csv(index=False)
    audit_df = pd.DataFrame(st.session_state.audit_log)
    audit_csv = audit_df.to_csv(index=False)

    control_df = (
        st.session_state.extraction_control_df
        if isinstance(st.session_state.extraction_control_df, pd.DataFrame)
        and not st.session_state.extraction_control_df.empty
        else build_extraction_control_df(st.session_state.df)
    )
    control_csv = control_df.to_csv(index=False)

    marker_payload = "|".join([
        current_hash,
        hashlib.sha256(review_csv.encode("utf-8")).hexdigest()[:16],
        hashlib.sha256(control_csv.encode("utf-8")).hexdigest()[:16],
        _template_cache_marker(),
        TOOL_VERSION,
    ])
    marker = hashlib.sha256(marker_payload.encode("utf-8")).hexdigest()[:20]

    try:
        if (
            st.session_state.get("working_predraft_excel_marker") != marker
            or not isinstance(st.session_state.get("working_predraft_excel_bytes"), (bytes, bytearray))
        ):
            with st.spinner("Preparing pre-draft Excel…"):
                base_bytes, excel_mime, excel_ext = _build_excel_bytes_cached(
                    df_csv,
                    True,
                    audit_csv,
                    review_csv,
                    control_csv,
                    _template_cache_marker(),
                )
                final_bytes = add_working_predraft_notice_sheet(base_bytes, excel_ext)
                st.session_state.working_predraft_excel_bytes = final_bytes
                st.session_state.working_predraft_excel_mime = excel_mime
                st.session_state.working_predraft_excel_ext = excel_ext
                st.session_state.working_predraft_excel_marker = marker

        def _safe_filename_part(value: str, fallback: str) -> str:
            clean = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in value)
            clean = clean.strip("_")
            return clean or fallback

        safe_site = _safe_filename_part(site, "Site")
        safe_product = _safe_filename_part(product, "Product")
        filename = (
            f"Pre-draft_{safe_site}_{safe_product}_{time.strftime('%Y%m%d_%H%M')}."
            f"{st.session_state.working_predraft_excel_ext}"
        )

        downloaded = st.download_button(
            label=f"📥 Download Pre-draft — {site} | {product}",
            data=st.session_state.working_predraft_excel_bytes,
            file_name=filename,
            mime=st.session_state.working_predraft_excel_mime,
            use_container_width=True,
            key="download_working_predraft_excel_step2",
        )

        if downloaded:
            audit(
                "predraft_excel_downloaded",
                f"file={filename}; reviewer={st.session_state.reviewer_name}; validated_hash={validated_hash}",
            )
            autosave_review_session("predraft_excel_downloaded")

        st.caption(
            "This is the pre-draft output. The final approved Excel remains available only after completion of the controlled review."
        )

    except Exception as e:
        st.error(f"Could not prepare the pre-draft Excel: {type(e).__name__}: {e}")
        st.code(traceback.format_exc())


def render_export_step():
    st.markdown('<div class="panel">', unsafe_allow_html=True)
    step_title("Export Excel", "Step 3: Export Excel file")

    msgs = []
    if needs_product():
        msgs.append("• Please enter **Product name** first (Step 1).")
    if needs_data():
        msgs.append("• Please **upload Batch Record (PDF/Word)** OR **upload Review Pack (Excel)** first (Step 1).")
    if msgs:
        st.warning("You can’t continue yet:\n\n" + "\n".join(msgs))
        st.markdown("</div>", unsafe_allow_html=True)
        return

    if st.session_state.review_status != "Approved":
        st.error("Export blocked: review status must be Approved.")
        st.info("Go to Step 2 · Review and click 'Approve & lock'.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    if not st.session_state.reviewer_name.strip():
        st.error("Export blocked: reviewer name is required.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    propagate_reviewer_to_df(fill_all_rows=False)

    df_csv = drop_internal_cols(st.session_state.df).to_csv(index=False)
    review_df = make_review_sheet_df()
    review_csv = review_df.to_csv(index=False)
    audit_df = pd.DataFrame(st.session_state.audit_log)
    audit_csv = audit_df.to_csv(index=False)

    control_df = (
        st.session_state.extraction_control_df
        if isinstance(st.session_state.extraction_control_df, pd.DataFrame)
        and not st.session_state.extraction_control_df.empty
        else build_extraction_control_df(st.session_state.df)
    )
    control_csv = control_df.to_csv(index=False)

    with st.spinner("Preparing Excel…"):
        excel_bytes, excel_mime, excel_ext = _build_excel_bytes_cached(
            df_csv, True, audit_csv, review_csv, control_csv, _template_cache_marker()
        )

    filename = f"{st.session_state.site}_FINAL_{time.strftime('%Y%m%d_%H%M')}.{excel_ext}"

    if excel_ext == EXT_XLSM:
        st.success(
            f"Ready to export ✅  Reviewer: {st.session_state.reviewer_name} · "
            f"Reviewed at: {st.session_state.reviewed_at}  "
            f"(Macro-enabled: double-click 'Reviewed & OK' to toggle and stamp the date.)"
        )
    else:
        st.success(
            f"Ready to export ✅  Reviewer: {st.session_state.reviewer_name} · "
            f"Reviewed at: {st.session_state.reviewed_at}"
        )
        st.info("template.xlsm not found — exporting plain .xlsx without macros.")

    st.download_button(
        label="📥 Download FINAL Excel",
        data=excel_bytes,
        file_name=filename,
        mime=excel_mime,
        use_container_width=True,
    )
    audit("final_excel_download_ready", filename)

    st.markdown("</div>", unsafe_allow_html=True)


# =============================================================================
# Router
# =============================================================================
if st.session_state.view_step == "Setup":
    render_setup_step()
elif st.session_state.view_step == "Review":
    render_review_step()
elif st.session_state.view_step == "Export Excel":
    render_export_step()

st.markdown(f"<div class='footer-version'>Tool version: {TOOL_VERSION}</div>", unsafe_allow_html=True)
